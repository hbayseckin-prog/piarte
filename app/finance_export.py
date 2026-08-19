"""Finans raporları — Excel (openpyxl) ve PDF (fpdf2) dışa aktarma."""
from __future__ import annotations

import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence
from urllib.parse import urlencode

from fastapi.responses import Response, StreamingResponse

_ROOT = Path(__file__).resolve().parent.parent
_BUNDLED_REGULAR = _ROOT / "fonts" / "DejaVuSans.ttf"
_BUNDLED_BOLD = _ROOT / "fonts" / "DejaVuSans-Bold.ttf"


def method_ui_label(method: str | None) -> str:
	m = (method or "").strip()
	if m == "EFT":
		return "IBAN"
	return m or "—"


def build_export_qs(**params: Any) -> str:
	cleaned = {k: v for k, v in params.items() if v not in (None, "", False)}
	for k, v in list(cleaned.items()):
		if v is True:
			cleaned[k] = "1"
	return urlencode({k: str(v) for k, v in cleaned.items()})


def excel_response(
	*,
	filename: str,
	sheets: Sequence[tuple[str, Sequence[str], Sequence[Sequence[Any]]]],
) -> StreamingResponse:
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill

	wb = Workbook()
	wb.remove(wb.active)
	header_fill = PatternFill("solid", fgColor="E0F2FE")
	header_font = Font(bold=True, color="0369A1")

	for sheet_name, headers, rows in sheets:
		ws = wb.create_sheet(title=(sheet_name or "Sayfa")[:31])
		ws.append(list(headers))
		for cell in ws[1]:
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = Alignment(horizontal="center")
		for row in rows:
			ws.append([_excel_cell(v) for v in row])
		for col in ws.columns:
			max_len = 0
			col_letter = col[0].column_letter
			for cell in col:
				max_len = max(max_len, len(str(cell.value or "")))
			ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)

	buf = BytesIO()
	wb.save(buf)
	buf.seek(0)
	safe = filename if filename.endswith(".xlsx") else f"{filename}.xlsx"
	return StreamingResponse(
		buf,
		media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		headers={"Content-Disposition": f'attachment; filename="{safe}"'},
	)


def _excel_cell(value: Any) -> Any:
	if value is None:
		return ""
	if isinstance(value, float):
		return round(value, 2)
	return value


def _find_unicode_fonts() -> tuple[str | None, str | None]:
	"""(regular_ttf, bold_ttf). Önce repodaki fonts/, sonra sistem."""
	regular_candidates = [
		str(_BUNDLED_REGULAR),
		"/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
		"/usr/share/fonts/TTF/DejaVuSans.ttf",
		"/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
		"C:/Windows/Fonts/arial.ttf",
		os.path.join(os.environ.get("WINDIR", "C:/Windows"), "Fonts", "arial.ttf"),
	]
	bold_candidates = [
		str(_BUNDLED_BOLD),
		"/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
		"/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
		"C:/Windows/Fonts/arialbd.ttf",
		os.path.join(os.environ.get("WINDIR", "C:/Windows"), "Fonts", "arialbd.ttf"),
	]
	regular = next((p for p in regular_candidates if p and os.path.isfile(p)), None)
	bold = next((p for p in bold_candidates if p and os.path.isfile(p)), None)
	if regular and not bold:
		bold = regular
	return regular, bold


def _latin1_safe(text: str) -> str:
	"""Helvetica fallback için Türkçe karakterleri sadeleştir."""
	repl = {
		"ç": "c", "Ç": "C", "ğ": "g", "Ğ": "G", "ı": "i", "İ": "I",
		"ö": "o", "Ö": "O", "ş": "s", "Ş": "S", "ü": "u", "Ü": "U",
		"₺": "TL", "→": "-", "—": "-", "·": "-",
	}
	out = "".join(repl.get(ch, ch) for ch in text)
	return out.encode("latin-1", errors="replace").decode("latin-1")


def pdf_response(
	*,
	filename: str,
	title: str,
	meta_lines: Sequence[str],
	headers: Sequence[str],
	rows: Sequence[Sequence[Any]],
	footer_note: str | None = None,
) -> Response:
	from fpdf import FPDF

	pdf = FPDF(orientation="L", unit="mm", format="A4")
	pdf.set_auto_page_break(auto=True, margin=12)
	pdf.add_page()

	regular_path, bold_path = _find_unicode_fonts()
	unicode_ok = bool(regular_path)
	if unicode_ok:
		pdf.add_font("ReportFont", "", regular_path)
		pdf.add_font("ReportFont", "B", bold_path or regular_path)
		font = "ReportFont"

		def _t(val: Any) -> str:
			if val is None:
				return ""
			if isinstance(val, float):
				return f"{val:.2f}"
			return str(val)
	else:
		font = "Helvetica"

		def _t(val: Any) -> str:
			if val is None:
				return ""
			if isinstance(val, float):
				return f"{val:.2f}"
			return _latin1_safe(str(val))

	pdf.set_font(font, "B", 14)
	pdf.cell(0, 8, _t(title), new_x="LMARGIN", new_y="NEXT")
	pdf.set_font(font, "", 9)
	pdf.set_text_color(80, 80, 80)
	for line in meta_lines:
		pdf.cell(0, 5, _t(line), new_x="LMARGIN", new_y="NEXT")
	pdf.set_text_color(0, 0, 0)
	pdf.ln(3)

	usable = pdf.w - pdf.l_margin - pdf.r_margin
	col_count = max(len(headers), 1)
	widths = [usable / col_count] * col_count
	if col_count >= 4:
		widths[0] = usable * 0.22
		rest = (usable - widths[0]) / (col_count - 1)
		for i in range(1, col_count):
			widths[i] = rest

	row_h = 6
	pdf.set_font(font, "B", 8)
	pdf.set_fill_color(224, 242, 254)
	for i, h in enumerate(headers):
		pdf.cell(widths[i], row_h, _t(h)[:40], border=1, fill=True)
	pdf.ln(row_h)

	pdf.set_font(font, "", 8)
	for row in rows:
		if pdf.get_y() > pdf.h - 18:
			pdf.add_page()
			pdf.set_font(font, "B", 8)
			pdf.set_fill_color(224, 242, 254)
			for i, h in enumerate(headers):
				pdf.cell(widths[i], row_h, _t(h)[:40], border=1, fill=True)
			pdf.ln(row_h)
			pdf.set_font(font, "", 8)
		vals = list(row) + [""] * max(0, col_count - len(row))
		for i in range(col_count):
			pdf.cell(widths[i], row_h, _t(vals[i])[:60], border=1)
		pdf.ln(row_h)

	if footer_note:
		pdf.ln(4)
		pdf.set_font(font, "", 8)
		pdf.set_text_color(100, 100, 100)
		pdf.cell(0, 5, _t(footer_note), new_x="LMARGIN", new_y="NEXT")

	pdf.set_y(-10)
	pdf.set_font(font, "", 7)
	pdf.set_text_color(120, 120, 120)
	pdf.cell(0, 4, _t(f"Piarte Finans · {datetime.now().strftime('%d.%m.%Y %H:%M')}"), align="R")

	out = bytes(pdf.output())
	safe = filename if filename.endswith(".pdf") else f"{filename}.pdf"
	return Response(
		content=out,
		media_type="application/pdf",
		headers={"Content-Disposition": f'attachment; filename="{safe}"'},
	)


def period_label(start: str | None, end: str | None) -> str:
	return f"Dönem: {start or '—'} → {end or '—'}"
