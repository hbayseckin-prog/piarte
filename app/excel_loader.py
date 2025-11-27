from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache
from pathlib import Path
import re
import unicodedata
from typing import Iterable, List, Optional

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT_DIR / "durum.xlsx"


@dataclass(frozen=True)
class RosterRow:
    student: str
    guardian: str | None
    phone: str | None


@dataclass(frozen=True)
class TeacherRoster:
    sheet_title: str
    teacher_display: str
    search_key: str
    rows: List[RosterRow]


@dataclass(frozen=True)
class DurumDataset:
    updated_at: datetime | None
    rosters: List[TeacherRoster]


def get_durum_dataset() -> DurumDataset:
    if not EXCEL_PATH.exists():
        return DurumDataset(updated_at=None, rosters=[])
    cache_key = EXCEL_PATH.stat().st_mtime
    return _load_dataset(cache_key)


def get_roster_for_teacher(name: str | None) -> Optional[TeacherRoster]:
    if not name:
        return None
    dataset = get_durum_dataset()
    target = _normalize_key(name)
    if not target:
        return None
    for roster in dataset.rosters:
        if target in roster.search_key or roster.search_key in target:
            return roster
    # fallback: partial token match
    for roster in dataset.rosters:
        if _tokens_overlap(target, roster.search_key):
            return roster
    return None


def _tokens_overlap(target: str, candidate: str) -> bool:
    target_tokens = _split_tokens(target)
    candidate_tokens = _split_tokens(candidate)
    return bool(target_tokens & candidate_tokens)


def _split_tokens(value: str) -> set[str]:
    if not value:
        return set()
    tokens = re.findall(r"[A-Z0-9]+", value)
    return set(tokens)


@lru_cache(maxsize=1)
def _load_dataset(cache_key: float) -> DurumDataset:  # pragma: no cover - IO helper
    wb = load_workbook(EXCEL_PATH, data_only=True)
    rosters: list[TeacherRoster] = []
    for ws in wb.worksheets:
        rows = list(_iter_rows(ws.iter_rows(values_only=True)))
        if not rows:
            continue
        sheet_title = ws.title.strip()
        roster = TeacherRoster(
            sheet_title=sheet_title,
            teacher_display=_derive_teacher_display(sheet_title),
            search_key=_normalize_key(sheet_title),
            rows=rows,
        )
        rosters.append(roster)
    updated_at = datetime.fromtimestamp(cache_key)
    return DurumDataset(updated_at=updated_at, rosters=rosters)


def _iter_rows(rows: Iterable[tuple]) -> Iterable[RosterRow]:
    seen_header = False
    for raw_row in rows:
        if not seen_header:
            seen_header = True
            continue
        student = _clean_str(raw_row, 0)
        guardian = _clean_str(raw_row, 1)
        phone = _format_phone(_get_value(raw_row, 2))
        if not student:
            continue
        yield RosterRow(student=student, guardian=guardian, phone=phone)


def _clean_str(row: tuple, idx: int) -> str | None:
    value = _get_value(row, idx)
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
    else:
        value = str(value).strip()
    if not value or value in {"?", "-"}:
        return None
    return value


def _get_value(row: tuple, idx: int):
    if len(row) <= idx:
        return None
    return row[idx]


def _format_phone(value) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    if not digits:
        return None
    if len(digits) == 10:
        digits = f"0{digits}"
    if len(digits) > 11:
        digits = digits[-11:]
    if len(digits) != 11:
        return digits
    return f"{digits[:4]} {digits[4:7]} {digits[7:9]} {digits[9:]}"


def _derive_teacher_display(title: str) -> str:
    cleaned = re.sub(r"(?i)t[üu]m|yoklama|hoca|hoc[ae]", "", title)
    cleaned = cleaned.replace(".", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip("- _")
    return cleaned.strip()


def _normalize_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_text = normalized.encode("ascii", "ignore").decode().upper()
    return re.sub(r"[^A-Z0-9]", "", ascii_text)





