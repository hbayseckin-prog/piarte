from fastapi import FastAPI, Depends, HTTPException, Request, Form, status
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse
from fastapi import Response
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, select
import os

from .db import Base, engine, get_db
from . import crud, schemas, models
try:
	from . import push_notify
except ImportError:
	push_notify = None
try:
    from . import excel_sync
except ImportError:
    excel_sync = None
try:
    from .seed import seed_courses, seed_admin
except ImportError:
    seed_courses = None
    seed_admin = None


def redirect_teacher(user):
    if user and user.get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    return None


def default_panel_url(user: dict | None) -> str:
    if not user:
        return "/"
    role = user.get("role")
    if role == "teacher":
        return "/ui/teacher"
    if role == "staff":
        return "/ui/staff"
    return "/dashboard"


def safe_return_url(url: str | None, default: str) -> str:
    if not url or not str(url).strip():
        return default
    path = str(url).strip()
    if not path.startswith("/") or path.startswith("//") or "://" in path:
        return default
    return path


def set_flash_success(request: Request, message: str) -> None:
    request.session["flash_success"] = message


def attendance_new_url(lesson_id: int, return_to: str | None = None, **query) -> str:
    from urllib.parse import urlencode
    params = {k: v for k, v in query.items() if v is not None and v != ""}
    if return_to:
        params["return_to"] = return_to
    qs = urlencode(params)
    return f"/lessons/{lesson_id}/attendance/new" + (f"?{qs}" if qs else "")


def calculate_next_lesson_date(original_date):
    """
    Haftalık tekrarlanan dersler için bugünden sonraki ilgili günü hesaplar.
    Örneğin: Orijinal tarih Pazartesi ise, bugün Pazartesi ise bugünü, 
    değilse bugünden sonraki Pazartesi'yi döndürür.
    
    Args:
        original_date: Orijinal ders tarihi (date objesi)
    
    Returns:
        Bugün veya bugünden sonraki ilgili günün tarihi (date objesi)
    """
    from datetime import date, timedelta
    
    today = date.today()
    original_weekday = original_date.weekday()  # 0=Pazartesi, 6=Pazar
    today_weekday = today.weekday()
    
    # Bugünden sonraki ilgili günü bul
    days_ahead = original_weekday - today_weekday
    if days_ahead < 0:  # Bu hafta geçtiyse gelecek hafta
        days_ahead += 7
    # days_ahead == 0 ise bugün o gün, bugünü döndür
    
    next_date = today + timedelta(days=days_ahead)
    return next_date


def filter_students_by_passive_flag(students, show_passive_students: bool):
	"""Ders programlarında pasif öğrenciler gösterilmez; yoklama/puantaj listeleri için show_passive_students=True kullanılabilir."""
	if show_passive_students:
		return students
	return [s for s in students if getattr(s, "is_active", True)]


WEEKDAY_NAMES = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]


def dedupe_daily_students_in_schedule(entries: list[dict]) -> list[dict]:
    """
    Aynı öğrenci aynı gün birden fazla ders slotunda görünüyorsa tek slotta bırakır.
    Öncelik: daha geç başlangıç saati, saat eşitse daha yeni lesson id.
    """
    chosen: dict[tuple[str, int], dict] = {}
    for entry in entries:
        lesson = entry.get("lesson")
        weekday = entry.get("weekday", "")
        students = entry.get("students") or []
        start_time = getattr(lesson, "start_time", None)
        lesson_id = getattr(lesson, "id", 0) or 0
        start_sort = (start_time.hour, start_time.minute) if start_time else (-1, -1)
        for s in students:
            key = (weekday, getattr(s, "id", 0))
            prev = chosen.get(key)
            if not prev:
                chosen[key] = {"entry": entry, "start_sort": start_sort, "lesson_id": lesson_id}
                continue
            if start_sort > prev["start_sort"] or (start_sort == prev["start_sort"] and lesson_id > prev["lesson_id"]):
                chosen[key] = {"entry": entry, "start_sort": start_sort, "lesson_id": lesson_id}

    filtered_entries: list[dict] = []
    for entry in entries:
        weekday = entry.get("weekday", "")
        kept_students = []
        for s in entry.get("students") or []:
            key = (weekday, getattr(s, "id", 0))
            if chosen.get(key, {}).get("entry") is entry:
                kept_students.append(s)
        if kept_students:
            new_entry = dict(entry)
            new_entry["students"] = kept_students
            filtered_entries.append(new_entry)
    return filtered_entries


def format_lessons_for_schedule(entries: list[dict]) -> list[dict]:
    formatted_lessons = []
    for entry in entries:
        lesson = entry["lesson"]
        students_for_view = filter_students_by_passive_flag(entry["students"], False)
        if not students_for_view:
            continue
        weekday = WEEKDAY_NAMES[lesson.lesson_date.weekday()] if hasattr(lesson.lesson_date, "weekday") else ""
        current_lesson_date = calculate_next_lesson_date(lesson.lesson_date)
        formatted_lessons.append({
            "weekday": weekday,
            "lesson": lesson,
            "current_lesson_date": current_lesson_date,
            "students": students_for_view,
        })
    return dedupe_daily_students_in_schedule(formatted_lessons)


def build_teachers_schedules(db: Session, teachers: list) -> list[dict]:
    lessons_by_teacher = crud.lessons_with_students_by_teacher_ids(db, [teacher.id for teacher in teachers])
    teachers_schedules = []
    for teacher in teachers:
        formatted_lessons = format_lessons_for_schedule(lessons_by_teacher.get(teacher.id, []))
        teachers_schedules.append({
            "teacher": teacher,
            "lessons": formatted_lessons,
        })
    return teachers_schedules


# Alt klasör desteği için root_path (eğer /piarte altında çalışıyorsa)
# Production'da environment variable veya Nginx yapılandırması ile ayarlanabilir
ROOT_PATH = os.getenv("ROOT_PATH", "")  # Varsayılan: boş (root'ta çalışır)

app = FastAPI(title="Piarte Kurs Yönetimi", root_path=ROOT_PATH)

# Uygulama başlangıcında migration kontrolü
@app.on_event("startup")
async def startup_event():
	"""Uygulama başlangıcında hafif migration kontrolü"""
	import logging
	try:
		from app.db import (
			ensure_is_active_column,
			ensure_teacher_is_active_column,
			ensure_teacher_hourly_rate_column,
			ensure_lesson_students_backfill_from_attendance,
			ensure_expenses_table,
			engine,
			Base,
		)
		ensure_is_active_column()
		ensure_teacher_is_active_column()
		ensure_teacher_hourly_rate_column()
		ensure_lesson_students_backfill_from_attendance()
		ensure_expenses_table()
		# Yeni Expense tablosu için metadata create (mevcut tablolara dokunmaz)
		Base.metadata.create_all(bind=engine, tables=[models.Expense.__table__])
		if push_notify:
			push_notify.ensure_push_subscriptions_table()
			push_notify.ensure_vapid_meta_table()
			push_notify.get_vapid_keys()
			Base.metadata.create_all(bind=engine, tables=[models.PushSubscription.__table__])
	except Exception as e:
		logging.error(f"Startup migration hatasi: {e}")

# CORS ayarları - iframe ve farklı domain'den erişim için
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Production'da belirli domain'ler belirtin: ["https://example.com"]
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Session secret key - environment variable'dan al, yoksa varsayılan kullan
SECRET_KEY = os.getenv("SECRET_KEY", "change-this-secret-key-in-production")
IS_PRODUCTION = os.getenv("RAILWAY_ENVIRONMENT") == "production" or os.getenv("ENVIRONMENT") == "production"
app.add_middleware(
	SessionMiddleware,
	secret_key=SECRET_KEY,
	same_site="lax",
	https_only=IS_PRODUCTION,
	max_age=14 * 24 * 60 * 60,
)
try:
	from uvicorn.middleware.proxy_headers import ProxyHeadersMiddleware
	app.add_middleware(ProxyHeadersMiddleware, trusted_hosts="*")
except ImportError:
	pass
templates = Jinja2Templates(directory="templates")

# Static files için - logo ve diğer statik dosyalar (proje root dizini)
# Logo dosyası root dizininde olduğu için root'u mount ediyoruz
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PWA_ICON_FILES = frozenset({
    "apple-touch-icon.png",
    "piarte-icon-192.png",
    "piarte-icon-512.png",
    "piarte-icon-512-maskable.png",
})
if os.path.exists(base_dir):
    app.mount("/static", StaticFiles(directory=base_dir), name="static")


@app.get("/icons/{filename}", name="pwa_icon", include_in_schema=False)
def pwa_icon(filename: str):
    if filename not in PWA_ICON_FILES:
        raise HTTPException(status_code=404)
    icon_path = os.path.join(base_dir, "icons", filename)
    if not os.path.exists(icon_path):
        raise HTTPException(status_code=404)
    return FileResponse(
        icon_path,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=604800"},
    )


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    icon_path = os.path.join(base_dir, "icons", "piarte-icon-192.png")
    if os.path.exists(icon_path):
        return FileResponse(icon_path, media_type="image/png")
    logo_path = os.path.join(base_dir, "piarte_logo.jpg")
    if os.path.exists(logo_path):
        return FileResponse(logo_path, media_type="image/jpeg")
    raise HTTPException(status_code=404)


@app.get("/manifest.webmanifest", include_in_schema=False)
def web_manifest():
    return JSONResponse(
        {
            "id": "/",
            "name": "Piarte",
            "short_name": "Piarte",
            "description": "Piarte okul yonetim paneli",
            "start_url": "/dashboard",
            "scope": "/",
            "display": "standalone",
            "background_color": "#ffffff",
            "theme_color": "#0ea5e9",
            "icons": [
                {
                    "src": "/icons/piarte-icon-192.png",
                    "sizes": "192x192",
                    "type": "image/png",
                    "purpose": "any",
                },
                {
                    "src": "/icons/piarte-icon-512.png",
                    "sizes": "512x512",
                    "type": "image/png",
                    "purpose": "any",
                },
                {
                    "src": "/icons/piarte-icon-512-maskable.png",
                    "sizes": "512x512",
                    "type": "image/png",
                    "purpose": "maskable",
                },
            ],
        },
        media_type="application/manifest+json",
        headers={"Cache-Control": "public, max-age=3600"},
    )


@app.get("/sw.js", include_in_schema=False)
def service_worker():
    sw_path = os.path.join(base_dir, "sw.js")
    if not os.path.exists(sw_path):
        raise HTTPException(status_code=404)
    return FileResponse(
        sw_path,
        media_type="application/javascript; charset=utf-8",
        headers={
            "Cache-Control": "no-cache",
            "Service-Worker-Allowed": "/",
        },
    )


@app.get("/api/push/vapid-public-key")
def api_push_vapid_public_key(request: Request):
    require_admin(request)
    if not push_notify:
        raise HTTPException(status_code=503, detail="Push desteklenmiyor")
    key = push_notify.get_vapid_public_key()
    if not key:
        raise HTTPException(status_code=503, detail="VAPID anahtarı yok")
    return {"publicKey": key}


@app.post("/api/push/subscribe")
async def api_push_subscribe(request: Request, db: Session = Depends(get_db)):
    user = require_admin(request)
    if not push_notify:
        raise HTTPException(status_code=503, detail="Push desteklenmiyor")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Geçersiz JSON")
    endpoint = (body.get("endpoint") or "").strip()
    keys = body.get("keys") or {}
    p256dh = (keys.get("p256dh") or "").strip()
    auth = (keys.get("auth") or "").strip()
    if not endpoint or not p256dh or not auth:
        raise HTTPException(status_code=400, detail="Eksik abonelik bilgisi")
    push_notify.upsert_subscription(
        db,
        user_id=int(user["id"]),
        endpoint=endpoint,
        p256dh=p256dh,
        auth=auth,
        user_agent=request.headers.get("user-agent"),
    )
    return {"ok": True}


@app.post("/api/push/unsubscribe")
async def api_push_unsubscribe(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    if not push_notify:
        raise HTTPException(status_code=503, detail="Push desteklenmiyor")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Geçersiz JSON")
    endpoint = (body.get("endpoint") or "").strip()
    if not endpoint:
        raise HTTPException(status_code=400, detail="endpoint gerekli")
    push_notify.delete_subscription_by_endpoint(db, endpoint)
    return {"ok": True}


@app.get("/api/push/status")
def api_push_status(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    if not push_notify:
        return {"ok": False, "subscriptions": 0, "vapid_sub": None}
    claims = push_notify.vapid_claims()
    return {
        "ok": True,
        "subscriptions": push_notify.count_admin_subscriptions(db),
        "vapid_sub": claims.get("sub"),
        "has_vapid": bool(push_notify.get_vapid_public_key()),
    }


@app.post("/api/push/test")
def api_push_test(request: Request, db: Session = Depends(get_db)):
    """Admin: anında test bildirimi gönder."""
    user = require_admin(request)
    if not push_notify:
        raise HTTPException(status_code=503, detail="Push desteklenmiyor")
    count = push_notify.count_admin_subscriptions(db)
    if count == 0:
        raise HTTPException(
            status_code=400,
            detail="Kayıtlı cihaz yok. Ana Ekran PWA'dan Bildirimleri açın.",
        )
    name = (user.get("full_name") or user.get("username") or "Admin").strip()
    result = push_notify.notify_admins_staff_cash(
        student_name="Test bildirimi",
        amount_try=0.0,
        staff_name=name,
        payment_id=None,
    )
    ok = result.get("sent", 0) > 0
    return {
        "ok": ok,
        "subscriptions": result.get("subscriptions", count),
        "sent": result.get("sent", 0),
        "failed": result.get("failed", 0),
        "vapid_sub": result.get("vapid_sub"),
        "errors": (result.get("errors") or [])[:3],
        "message": (
            f"{result.get('sent', 0)} cihaz başarılı"
            if ok
            else "Gönderilemedi: " + "; ".join((result.get("errors") or ["bilinmeyen hata"])[:2])
        ),
    }


# iframe güvenlik header'ları için middleware
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
	response = await call_next(request)
	# iframe'de çalışması için - SAMEORIGIN: aynı domain'den iframe'e izin verir
	# Tüm origin'ler için izin vermek isterseniz bu satırı kaldırın
	response.headers["X-Frame-Options"] = "SAMEORIGIN"
	# Güvenlik için
	response.headers["X-Content-Type-Options"] = "nosniff"
	return response

# Basit health check endpoint
@app.get("/health")
def health_check(db: Session = Depends(get_db)):
	from sqlalchemy import text
	try:
		db.execute(text("SELECT 1"))
		user_count = db.scalar(select(func.count(models.User.id))) or 0
		return {
			"status": "ok",
			"message": "Server is running",
			"database": "connected",
			"users": user_count,
		}
	except Exception as e:
		return {
			"status": "degraded",
			"message": "Server is running but database unreachable",
			"database": "error",
			"detail": str(e),
		}

# Veritabanı kurulum endpoint'i
@app.get("/setup-database", response_class=HTMLResponse)
def setup_database_endpoint(request: Request):
	"""Veritabanını oluştur ve seed data ekle - HTML response ile"""
	try:
		reset_performed = False
		try:
			# Tüm tabloları oluştur
			Base.metadata.create_all(bind=engine)
		except Exception as e:
			# Eğer DuplicateTable / already exists hatası ise tüm tabloları silip yeniden oluştur
			msg = str(e)
			pgcode = getattr(getattr(e, "orig", None), "pgcode", "")
			if "DuplicateTable" in msg or "already exists" in msg or pgcode == "42P07":
				reset_performed = True
				Base.metadata.drop_all(bind=engine)
				Base.metadata.create_all(bind=engine)
			else:
				raise
		
		# Seed data ekle
		db = next(get_db())
		messages = []
		errors = []
		
		try:
			from app.seed import seed_courses, seed_admin
			
			# Kursları ekle
			if seed_courses:
				try:
					seed_courses(db)
					messages.append("✅ Kurslar başarıyla eklendi")
				except Exception as e:
					errors.append(f"⚠️ Kurs ekleme hatası: {str(e)}")
			
			# Admin kullanıcısını ekle
			if seed_admin:
				try:
					seed_admin(db)
					messages.append("✅ Admin kullanıcısı eklendi (kullanıcı adı: admin, şifre: admin123)")
				except Exception as e:
					errors.append(f"⚠️ Admin ekleme hatası: {str(e)}")
			
			db.commit()
		except Exception as e:
			errors.append(f"❌ Seed data hatası: {str(e)}")
			db.rollback()
		finally:
			db.close()
		
		# HTML response oluştur
		if reset_performed:
			messages.insert(0, "ℹ️ Mevcut tablolar silinip yeniden oluşturuldu (duplicate hata nedeniyle).")
		messages_html = "\n".join([f"<p style='color: green;'>{msg}</p>" for msg in messages])
		errors_html = "\n".join([f"<p style='color: orange;'>{err}</p>" for err in errors])
		
		html_content = f"""
		<!DOCTYPE html>
		<html lang="tr">
		<head>
			<meta charset="UTF-8">
			<meta name="viewport" content="width=device-width, initial-scale=1.0">
			<title>Veritabanı Kurulumu - Piarte</title>
			<style>
				body {{
					font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
					background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
					margin: 0;
					padding: 20px;
					min-height: 100vh;
					display: flex;
					justify-content: center;
					align-items: center;
				}}
				.container {{
					background: white;
					border-radius: 15px;
					padding: 40px;
					max-width: 600px;
					box-shadow: 0 10px 40px rgba(0,0,0,0.2);
				}}
				h1 {{
					color: #667eea;
					margin-bottom: 20px;
					text-align: center;
				}}
				.status {{
					background: #f0f9ff;
					border-left: 4px solid #667eea;
					padding: 15px;
					margin: 20px 0;
					border-radius: 5px;
				}}
				.success {{
					background: #f0fdf4;
					border-left: 4px solid #22c55e;
				}}
				.warning {{
					background: #fffbeb;
					border-left: 4px solid #f59e0b;
				}}
				.error {{
					background: #fef2f2;
					border-left: 4px solid #ef4444;
				}}
				.button {{
					display: inline-block;
					background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
					color: white;
					padding: 12px 30px;
					text-decoration: none;
					border-radius: 5px;
					margin-top: 20px;
					text-align: center;
					width: 100%;
					box-sizing: border-box;
				}}
				.button:hover {{
					opacity: 0.9;
				}}
				.info {{
					background: #f8fafc;
					padding: 15px;
					border-radius: 5px;
					margin-top: 20px;
					font-size: 14px;
					color: #64748b;
				}}
			</style>
		</head>
		<body>
			<div class="container">
				<h1>📦 Veritabanı Kurulumu</h1>
				
				<div class="status success">
					<strong>✅ Tablolar başarıyla oluşturuldu!</strong>
				</div>
				
				{messages_html if messages_html else ""}
				{errors_html if errors_html else ""}
				
				<div class="info">
					<strong>📝 Sonraki Adımlar:</strong><br>
					1. Ana sayfaya dönün ve giriş yapın<br>
					2. Admin kullanıcısı ile giriş yapın (kullanıcı adı: <strong>admin</strong>, şifre: <strong>admin123</strong>)<br>
					3. Güvenlik için şifrenizi değiştirin!
				</div>
				
				<a href="/" class="button">🏠 Ana Sayfaya Dön</a>
			</div>
		</body>
		</html>
		"""
		
		return HTMLResponse(content=html_content)
		
	except Exception as e:
		# Hata durumunda HTML response
		html_content = f"""
		<!DOCTYPE html>
		<html lang="tr">
		<head>
			<meta charset="UTF-8">
			<meta name="viewport" content="width=device-width, initial-scale=1.0">
			<title>Veritabanı Kurulum Hatası - Piarte</title>
			<style>
				body {{
					font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
					background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
					margin: 0;
					padding: 20px;
					min-height: 100vh;
					display: flex;
					justify-content: center;
					align-items: center;
				}}
				.container {{
					background: white;
					border-radius: 15px;
					padding: 40px;
					max-width: 600px;
					box-shadow: 0 10px 40px rgba(0,0,0,0.2);
				}}
				h1 {{
					color: #ef4444;
					margin-bottom: 20px;
					text-align: center;
				}}
				.error {{
					background: #fef2f2;
					border-left: 4px solid #ef4444;
					padding: 15px;
					margin: 20px 0;
					border-radius: 5px;
					color: #991b1b;
				}}
				.button {{
					display: inline-block;
					background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
					color: white;
					padding: 12px 30px;
					text-decoration: none;
					border-radius: 5px;
					margin-top: 20px;
					text-align: center;
					width: 100%;
					box-sizing: border-box;
				}}
			</style>
		</head>
		<body>
			<div class="container">
				<h1>❌ Veritabanı Kurulum Hatası</h1>
				
				<div class="error">
					<strong>Hata:</strong><br>
					{str(e)}
				</div>
				
				<div style="margin-top: 20px; color: #64748b; font-size: 14px;">
					<strong>Çözüm Önerileri:</strong><br>
					1. Railway'de DATABASE_URL değişkeninin doğru olduğundan emin olun<br>
					2. PostgreSQL servisinin çalıştığını kontrol edin<br>
					3. Railway'de "Deploy Logs" sekmesinden hata detaylarını kontrol edin
				</div>
				
				<a href="/" class="button">🏠 Ana Sayfaya Dön</a>
			</div>
		</body>
		</html>
		"""
		return HTMLResponse(content=html_content, status_code=500)

# Startup event'ini kaldırdık - lazy initialization kullanacağız
# İlk database isteğinde otomatik olarak tablolar oluşturulacak


def require_user(request: Request):
	user = request.session.get("user")
	if not user:
		raise HTTPException(status_code=401, detail="Giriş gerekiyor")
	return user


def require_admin(request: Request):
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=401, detail="Giriş gerekiyor")
    role = (user.get("role") or "").strip().lower()
    if role != "admin":
        raise HTTPException(status_code=403, detail="Yetki yok")
    return user


def build_session_user_payload(user: models.User, role_override: str | None = None) -> dict:
    """DB user nesnesini session payload formatına çevirir."""
    role_value = (role_override or user.role or "admin").strip().lower()
    if role_value not in {"admin", "staff", "teacher"}:
        role_value = "admin"
    return {
        "id": user.id,
        "username": user.username,
        "full_name": user.full_name,
        "role": role_value,
        "teacher_id": getattr(user, "teacher_id", None),
    }


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
	# Kullanıcı giriş yapmışsa dashboard'a, yoksa index.html'i göster
	user = request.session.get("user")
	if user:
		if user.get("role") == "teacher":
			return RedirectResponse(url="/ui/teacher", status_code=302)
		elif user.get("role") == "staff":
			return RedirectResponse(url="/ui/staff", status_code=302)
		else:
			return RedirectResponse(url="/dashboard", status_code=302)
	# index.html'i göster
	try:
		with open("index.html", "r", encoding="utf-8") as f:
			return HTMLResponse(content=f.read())
	except FileNotFoundError:
		# index.html yoksa login sayfasına yönlendir
		return RedirectResponse(url="/login/admin", status_code=302)



@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    try:
        from passlib.hash import pbkdf2_sha256
        user = crud.get_user_by_username(db, username)
        if not user:
            return RedirectResponse(url="/", status_code=302)
        try:
            password_valid = pbkdf2_sha256.verify(password, user.password_hash)
        except Exception as e:
            import logging
            logging.error(f"Şifre doğrulama hatası: {e}")
            return RedirectResponse(url="/", status_code=302)
        if not password_valid:
            return RedirectResponse(url="/", status_code=302)
        request.session["user"] = {
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "role": user.role or "admin",
            "teacher_id": getattr(user, 'teacher_id', None),
        }
        return RedirectResponse(url="/dashboard", status_code=302)
    except Exception as e:
        import logging
        import traceback
        logging.error(f"Login hatası: {e}")
        logging.error(traceback.format_exc())
        return RedirectResponse(url="/", status_code=302)

@app.get("/logout")
def logout(request: Request):
	# Kullanıcının rolünü al (session temizlenmeden önce)
	user = request.session.get("user")
	role = user.get("role") if user else None
	
	# Session'ı temizle
	request.session.clear()
	
	# Rolüne göre ilgili giriş sayfasına yönlendir
	if role == "teacher":
		return RedirectResponse(url="/login/teacher", status_code=302)
	elif role == "staff":
		return RedirectResponse(url="/login/staff", status_code=302)
	else:
		# admin veya diğer durumlar için admin giriş sayfasına yönlendir
		return RedirectResponse(url="/login/admin", status_code=302)


@app.get("/session/switch-user/{target_user_id}")
def session_switch_user(target_user_id: int, request: Request, db: Session = Depends(get_db)):
    """Admin kullanıcısının çıkış yapmadan staff/teacher kullanıcıya geçmesini sağlar."""
    current_user = request.session.get("user")
    if not current_user or current_user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=302)

    target_user = db.get(models.User, target_user_id)
    target_role = (getattr(target_user, "role", "") or "").strip().lower() if target_user else ""
    if not target_user or target_role not in {"staff", "teacher"}:
        return RedirectResponse(url="/dashboard", status_code=302)

    # İlk geçişte admin oturumunu sakla; geri dönüşte kullanılacak.
    if not request.session.get("admin_original_user"):
        request.session["admin_original_user"] = current_user

    request.session["user"] = build_session_user_payload(target_user, role_override=target_role)
    if target_role == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    return RedirectResponse(url="/ui/staff", status_code=302)


@app.get("/session/switch-back-admin")
def session_switch_back_admin(request: Request):
    """Kullanıcı geçişinden sonra admin oturumuna geri döner."""
    original_admin_user = request.session.get("admin_original_user")
    if original_admin_user:
        request.session["user"] = original_admin_user
        request.session.pop("admin_original_user", None)
        return RedirectResponse(url="/dashboard", status_code=302)

    active_user = request.session.get("user")
    if active_user and active_user.get("role") == "admin":
        return RedirectResponse(url="/dashboard", status_code=302)
    return RedirectResponse(url="/login/admin", status_code=302)


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    db: Session = Depends(get_db),
    teacher_id: str | None = None,
    student_id: str | None = None,
    course_id: str | None = None,
    status: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    order_by: str = "marked_at_desc",
    student_name: str | None = None,
    payment_day: str | None = None,
    payment_status_filter: str | None = None,
    attendance_view: str | None = None,
):
    user = request.session.get("user")
    if not user:
        return RedirectResponse(url="/login/admin", status_code=302)
    if user.get("role") == "staff":
        return RedirectResponse(url="/ui/staff", status_code=302)
    if user.get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    courses = crud.list_courses(db)
    students = crud.list_students(db)
    teachers = crud.list_teachers(db)
    # Aktif / pasif öğrenci sayıları
    active_students_count = sum(1 for s in students if getattr(s, "is_active", True))
    passive_students_count = sum(1 for s in students if hasattr(s, "is_active") and s.is_active is False)
    # Staff (personel) kullanıcılarını getir
    from sqlalchemy import select
    staff_users = db.scalars(select(models.User).where(models.User.role == "staff").order_by(models.User.created_at.desc())).all()
    admin_switch_targets: list[dict] = []
    staff_switch_user = db.scalars(
        select(models.User)
        .where(models.User.role == "staff")
        .order_by(models.User.created_at.asc())
    ).first()
    if staff_switch_user:
        admin_switch_targets.append({
            "user_id": staff_switch_user.id,
            "label": "Staff Paneline Geç",
        })

    from sqlalchemy import or_
    target_teacher_user = db.scalars(
        select(models.User)
        .join(models.Teacher, models.Teacher.id == models.User.teacher_id)
        .where(
            models.User.role == "teacher",
            or_(
                models.Teacher.first_name.ilike("%Gokhan%"),
                models.Teacher.first_name.ilike("%Gökhan%"),
            ),
            or_(
                models.Teacher.last_name.ilike("%Husunbeyi%"),
                models.Teacher.last_name.ilike("%Hüsünbeyi%"),
            ),
        )
        .order_by(models.User.created_at.asc())
    ).first()
    if not target_teacher_user:
        target_teacher_user = db.scalars(
            select(models.User)
            .where(models.User.role == "teacher")
            .order_by(models.User.created_at.asc())
        ).first()
    if target_teacher_user:
        teacher_display = "Öğretmen Paneline Geç"
        target_teacher = db.get(models.Teacher, target_teacher_user.teacher_id) if target_teacher_user.teacher_id else None
        if target_teacher and (target_teacher.first_name or target_teacher.last_name):
            teacher_display = f"{target_teacher.first_name or ''} {target_teacher.last_name or ''} Paneline Geç".strip()
        admin_switch_targets.append({
            "user_id": target_teacher_user.id,
            "label": teacher_display,
        })
    
    # Query parametrelerini integer'a çevir (boş string'leri None yap)
    teacher_id_int = None
    student_id_int = None
    course_id_int = None
    if teacher_id and teacher_id.strip():
        try:
            teacher_id_int = int(teacher_id)
        except (ValueError, TypeError):
            teacher_id_int = None
    if student_id and student_id.strip():
        try:
            student_id_int = int(student_id)
        except (ValueError, TypeError):
            student_id_int = None
    if course_id and course_id.strip():
        try:
            course_id_int = int(course_id)
        except (ValueError, TypeError):
            course_id_int = None
    # Tarih filtrelerini parse et
    from datetime import date, datetime
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            y, m, d = map(int, start_date.split("-"))
            start_date_obj = date(y, m, d)
        except Exception:
            pass
    if end_date:
        try:
            y, m, d = map(int, end_date.split("-"))
            end_date_obj = date(y, m, d)
        except Exception:
            pass
    
    # Filtrelerin olup olmadığını kontrol et
    has_filters = any([
        teacher_id_int is not None,
        student_id_int is not None,
        course_id_int is not None,
        status is not None and status.strip(),
        start_date_obj is not None,
        end_date_obj is not None,
        student_name is not None and student_name.strip(),
    ])
    
    # Eğer hiçbir filtre yoksa, boş liste döndür
    if not has_filters:
        attendances = []
    else:
        attendances = crud.list_all_attendances(
            db,
            teacher_id=teacher_id_int,
            student_id=student_id_int,
            course_id=course_id_int,
            status=status,
            start_date=start_date_obj,
            end_date=end_date_obj,
            order_by=order_by,
            limit=200,
        )
        if student_name and student_name.strip() and not student_id_int:
            filtered = []
            for a in attendances:
                stu = db.get(models.Student, a.student_id)
                if not stu:
                    continue
                full_name = f"{stu.first_name} {stu.last_name}"
                if crud.student_name_matches_prefix(full_name, student_name):
                    filtered.append(a)
            attendances = filtered
    
    # Yoklamaları ders ve öğrenci bilgileriyle birlikte hazırla
    attendances_with_details = []
    if attendances:
        lesson_ids = {att.lesson_id for att in attendances}
        student_ids = {att.student_id for att in attendances}
        lessons_map = {l.id: l for l in db.scalars(select(models.Lesson).where(models.Lesson.id.in_(lesson_ids))).all()} if lesson_ids else {}
        students_map = {s.id: s for s in db.scalars(select(models.Student).where(models.Student.id.in_(student_ids))).all()} if student_ids else {}
        teacher_ids = {l.teacher_id for l in lessons_map.values() if l.teacher_id}
        course_ids = {l.course_id for l in lessons_map.values() if l.course_id}
        teachers_map = {t.id: t for t in db.scalars(select(models.Teacher).where(models.Teacher.id.in_(teacher_ids))).all()} if teacher_ids else {}
        courses_map = {c.id: c for c in db.scalars(select(models.Course).where(models.Course.id.in_(course_ids))).all()} if course_ids else {}
        for att in attendances:
            lesson = lessons_map.get(att.lesson_id)
            student = students_map.get(att.student_id)
            teacher = teachers_map.get(lesson.teacher_id) if lesson and lesson.teacher_id else None
            course = courses_map.get(lesson.course_id) if lesson and lesson.course_id else None
            attendances_with_details.append({
                "attendance": att,
                "lesson": lesson,
                "student": student,
                "teacher": teacher,
                "course": course,
            })
    # Puantaj / öğretmen özeti (filtre uygulandığında)
    attendance_report = []
    attendance_totals_by_teacher = {}
    attendance_teacher_summary = []
    attendance_summary_grand_totals = None
    _show_puantaj = (attendance_view or "yoklama").strip() in ("both", "puantaj")
    if user.get("role") == "admin" and has_filters:
        filtered_report = crud.get_attendance_report_by_teacher(
            db,
            teacher_id=teacher_id_int,
            student_id=student_id_int,
            course_id=course_id_int,
            start_date=start_date_obj,
            end_date=end_date_obj,
            status=status,
            student_name=student_name if not student_id_int else None,
        )
        for teacher_report in filtered_report:
            students = teacher_report.get("students") or []
            if not students:
                continue
            totals = {
                "total_present": sum(s.get("present", 0) for s in students),
                "total_excused_absent": sum(s.get("excused_absent", 0) for s in students),
                "total_telafi": sum(s.get("telafi", 0) for s in students),
                "total_unexcused_absent": sum(s.get("unexcused_absent", 0) for s in students),
                "total_lessons": sum(s.get("total", 0) for s in students),
            }
            teacher = teacher_report["teacher"]
            attendance_totals_by_teacher[teacher.id] = totals
            attendance_teacher_summary.append({"teacher": teacher, "totals": totals})
        if attendance_teacher_summary:
            attendance_summary_grand_totals = {
                "total_present": sum(item["totals"]["total_present"] for item in attendance_teacher_summary),
                "total_excused_absent": sum(item["totals"]["total_excused_absent"] for item in attendance_teacher_summary),
                "total_telafi": sum(item["totals"]["total_telafi"] for item in attendance_teacher_summary),
                "total_unexcused_absent": sum(item["totals"]["total_unexcused_absent"] for item in attendance_teacher_summary),
                "total_lessons": sum(item["totals"]["total_lessons"] for item in attendance_teacher_summary),
            }
        if _show_puantaj:
            attendance_report = filtered_report
    
    # Tüm öğretmenler için haftalık ders programını hazırla (saat bazlı grid için)
    teachers_schedules = build_teachers_schedules(db, teachers)
    
    # Ödeme durumu listesi ve ders bilgileri (sadece admin için)
    students_needing_payment = []
    students_needing_payment_lessons = {}
    payment_status_list = []
    payment_status_filter_value = ""
    if user.get("role") == "admin":
        payment_status_filter_value = (payment_status_filter or "").strip().lower()
        if payment_status_filter_value not in crud.VALID_PAYMENT_STATUS_FILTERS:
            payment_status_filter_value = ""
    
    context = {
        "request": request,
        "courses": courses,
        "students": students,
        "teachers": teachers,
        "staff_users": staff_users,
        "active_students_count": active_students_count,
        "passive_students_count": passive_students_count,
        "attendances": attendances_with_details,
        "attendance_report": attendance_report,
        "attendance_totals_by_teacher": attendance_totals_by_teacher,
        "attendance_teacher_summary": attendance_teacher_summary,
        "attendance_summary_grand_totals": attendance_summary_grand_totals,
        "has_attendance_filters": has_filters,
        "teachers_schedules": teachers_schedules,
        "students_needing_payment": students_needing_payment,
        "students_needing_payment_lessons": students_needing_payment_lessons,
        "payment_status_list": payment_status_list,
        "user": user,
        "admin_switch_targets": admin_switch_targets,
        "filters": {
            "teacher_id": str(teacher_id_int) if teacher_id_int else "",
            "student_id": str(student_id_int) if student_id_int else "",
            "course_id": str(course_id_int) if course_id_int else "",
            "status": status or "",
            "start_date": start_date or "",
            "end_date": end_date or "",
            "order_by": order_by,
            "student_name": student_name or "",
            "payment_day": payment_day or "",
            "payment_status_filter": payment_status_filter_value,
            "attendance_view": (attendance_view or "yoklama").strip() or "yoklama",
        },
    }
    return templates.TemplateResponse("dashboard.html", context)


@app.get("/ui/payment-status/partial", response_class=HTMLResponse)
def payment_status_partial(
    request: Request,
    db: Session = Depends(get_db),
    payment_status_filter: str | None = None,
    payment_day: str | None = None,
    payment_day_filter: str | None = None,
    variant: str = "admin",
):
    user = request.session.get("user")
    if not user:
        return HTMLResponse("", status_code=401)

    variant = (variant or "admin").strip().lower()
    if variant == "staff":
        if user.get("role") not in ("admin", "staff"):
            return HTMLResponse("", status_code=403)
        day_value = payment_day_filter or payment_day
        template_name = "_payment_status_staff_results.html"
    else:
        if user.get("role") != "admin":
            return HTMLResponse("", status_code=403)
        day_value = payment_day or payment_day_filter
        template_name = "_payment_status_admin_results.html"

    status_filter = (payment_status_filter or "").strip().lower()
    if status_filter not in crud.VALID_PAYMENT_STATUS_FILTERS:
        return HTMLResponse(
            '<p style="color:#64748b;text-align:center;padding:20px;">Listelemek için yukarıdan bir ödeme durumu seçin.</p>'
        )

    payment_status_list, students_needing_payment_lessons = crud.build_payment_status_list(
        db,
        status_filter=status_filter,
        payment_day=day_value,
        include_staff_fields=(variant == "staff"),
    )

    return templates.TemplateResponse(
        template_name,
        {
            "request": request,
            "payment_status_list": payment_status_list,
            "students_needing_payment_lessons": students_needing_payment_lessons,
            "payment_status_filter": status_filter,
        },
    )


@app.get("/dashboard/export/excel")
def export_punctuality_excel(
    request: Request,
    db: Session = Depends(get_db),
    teacher_id: str | None = None,
    student_id: str | None = None,
    course_id: str | None = None,
    status: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    student_name: str | None = None,
):
    """Puantaj tablosunu Excel formatında export eder"""
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Yetki yok")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime, date
    from io import BytesIO
    
    # Filtreleri parse et
    teacher_id_int = None
    student_id_int = None
    course_id_int = None
    start_date_obj = None
    end_date_obj = None
    
    if teacher_id and teacher_id.strip():
        try:
            teacher_id_int = int(teacher_id)
        except (ValueError, TypeError):
            pass
    if student_id and student_id.strip():
        try:
            student_id_int = int(student_id)
        except (ValueError, TypeError):
            pass
    if course_id and course_id.strip():
        try:
            course_id_int = int(course_id)
        except (ValueError, TypeError):
            pass
    if start_date:
        try:
            y, m, d = map(int, start_date.split("-"))
            start_date_obj = date(y, m, d)
        except Exception:
            pass
    if end_date:
        try:
            y, m, d = map(int, end_date.split("-"))
            end_date_obj = date(y, m, d)
        except Exception:
            pass
    
    # Puantaj raporunu getir (dashboard filtreleriyle aynı)
    attendance_report = crud.get_attendance_report_by_teacher(
        db,
        teacher_id=teacher_id_int,
        student_id=student_id_int,
        course_id=course_id_int,
        start_date=start_date_obj,
        end_date=end_date_obj,
        status=status,
        student_name=student_name if not student_id_int else None,
    )
    
    # Excel workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Puantaj Raporu"
    
    # Stil tanımlamaları
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Başlık satırı
    row = 1
    ws.merge_cells(f'A{row}:G{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = f"Puantaj Raporu - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # Her öğretmen için ayrı bölüm
    for teacher_report in attendance_report:
        # Öğretmen başlığı
        ws.merge_cells(f'A{row}:G{row}')
        teacher_cell = ws[f'A{row}']
        teacher_cell.value = f"Öğretmen: {teacher_report['teacher'].first_name} {teacher_report['teacher'].last_name}"
        teacher_cell.font = Font(bold=True, size=12, color="001F2937")
        teacher_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        teacher_cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 1
        
        # Öğrenci verileri (rapor zaten filtrelenmiş)
        students_data = teacher_report['students']
        
        if not students_data:
            ws.merge_cells(f'A{row}:G{row}')
            no_data_cell = ws[f'A{row}']
            no_data_cell.value = "Bu öğretmen için filtre kriterlerine uygun veri bulunmuyor."
            no_data_cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 2
            continue
        
        # Tablo başlıkları
        headers = ["Öğrenci", "Geldi", "Haberli Gelmedi", "Telafi", "Habersiz Gelmedi", "Toplam Ders", "Yoklama Tarihleri"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border_style
        row += 1
        
        # Öğrenci verileri
        for student_data in students_data:
            # Öğrenci adı
            cell = ws.cell(row=row, column=1)
            cell.value = f"{student_data['student'].first_name} {student_data['student'].last_name}"
            cell.border = border_style
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Geldi
            cell = ws.cell(row=row, column=2)
            cell.value = student_data['present']
            cell.border = border_style
            cell.alignment = center_alignment
            cell.font = Font(color="0010B981", bold=True)  # RGB format
            
            # Haberli Gelmedi
            cell = ws.cell(row=row, column=3)
            cell.value = student_data['excused_absent']
            cell.border = border_style
            cell.alignment = center_alignment
            cell.font = Font(color="00F97316", bold=True)  # RGB format
            
            # Telafi
            cell = ws.cell(row=row, column=4)
            cell.value = student_data['telafi']
            cell.border = border_style
            cell.alignment = center_alignment
            cell.font = Font(color="008B5CF6", bold=True)  # RGB format
            
            # Habersiz Gelmedi
            cell = ws.cell(row=row, column=5)
            cell.value = student_data['unexcused_absent']
            cell.border = border_style
            cell.alignment = center_alignment
            cell.font = Font(color="00EF4444", bold=True)  # RGB format
            
            # Toplam Ders
            cell = ws.cell(row=row, column=6)
            cell.value = student_data['total']
            cell.border = border_style
            cell.alignment = center_alignment
            cell.font = Font(bold=True)
            
            # Yoklama Tarihleri
            cell = ws.cell(row=row, column=7)
            dates = student_data.get('dates', [])
            if dates:
                # Tarihleri sırala ve tekrar edenleri kaldır
                unique_dates = sorted(list(set(dates)))
                cell.value = ', '.join(unique_dates)
            else:
                cell.value = '-'
            cell.border = border_style
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = Font(size=10)
            
            row += 1
        
        # Öğretmen bölümü sonrası boş satır
        row += 1
    
    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 50  # Yoklama Tarihleri sütunu için geniş sütun
    
    # Excel dosyasını memory'de oluştur
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Dosya adı
    filename = f"puantaj_raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


# UI: Quick search
@app.get("/ui/search", response_class=HTMLResponse)
def quick_search(request: Request, q: str, db: Session = Depends(get_db)):
    user = request.session.get("user")
    if not user:
        return RedirectResponse(url="/", status_code=302)
    if user.get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)

    term = f"%{q.strip()}%"
    students = db.query(models.Student).filter(
        (models.Student.first_name.ilike(term)) | (models.Student.last_name.ilike(term))
    ).limit(20).all()
    teachers = db.query(models.Teacher).filter(
        models.Teacher.is_active == True,
        (models.Teacher.first_name.ilike(term)) | (models.Teacher.last_name.ilike(term))
    ).limit(20).all()
    courses = db.query(models.Course).filter(models.Course.name.ilike(term)).limit(20).all()

    if user.get("role") == "admin" and len(students) == 1 and not teachers and not courses:
        from urllib.parse import urlencode
        params = {"q": q.strip(), "return_to": "/dashboard"}
        return RedirectResponse(
            url=f"/ui/search/student/{students[0].id}?{urlencode(params)}",
            status_code=302,
        )

    return templates.TemplateResponse(
        "search_results.html",
        {
            "request": request,
            "q": q,
            "students": students,
            "teachers": teachers,
            "courses": courses,
            "is_admin": user.get("role") == "admin",
        },
    )


@app.get("/ui/search/student/{student_id}", response_class=HTMLResponse)
def search_student_actions(
    student_id: int,
    request: Request,
    q: str | None = None,
    return_to: str | None = None,
    db: Session = Depends(get_db),
):
    """Admin hızlı arama: öğrenci yoklama/ödeme işlemleri."""
    from sqlalchemy.orm import joinedload

    user = request.session.get("user")
    if not user:
        return RedirectResponse(url="/", status_code=302)
    if user.get("role") != "admin":
        return RedirectResponse(url="/dashboard", status_code=302)

    student = crud.get_student(db, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Öğrenci bulunamadı")

    resolved_return_to = safe_return_url(return_to, "/dashboard")
    page_return_to = f"/ui/search/student/{student_id}"
    if q and q.strip():
        from urllib.parse import urlencode
        page_return_to += "?" + urlencode({"q": q.strip(), "return_to": resolved_return_to})

    payments = crud.list_payments_by_student(db, student_id)

    attendances_raw = crud.list_all_attendances(
        db, student_id=student_id, limit=500, order_by="marked_at_desc"
    )
    attendance_rows = []
    for att in attendances_raw:
        lesson = db.get(models.Lesson, att.lesson_id) if att.lesson_id else None
        course = db.get(models.Course, lesson.course_id) if lesson and lesson.course_id else None
        teacher = db.get(models.Teacher, lesson.teacher_id) if lesson and lesson.teacher_id else None
        attendance_rows.append({
            "attendance": att,
            "lesson": lesson,
            "course": course,
            "teacher": teacher,
        })

    lessons = db.scalars(
        select(models.Lesson)
        .join(models.LessonStudent, models.LessonStudent.lesson_id == models.Lesson.id)
        .where(models.LessonStudent.student_id == student_id)
        .options(joinedload(models.Lesson.course), joinedload(models.Lesson.teacher))
        .order_by(models.Lesson.lesson_date.asc(), models.Lesson.start_time.asc())
    ).all()

    return templates.TemplateResponse(
        "search_student_actions.html",
        {
            "request": request,
            "student": student,
            "q": q or "",
            "return_to": resolved_return_to,
            "page_return_to": page_return_to,
            "attendance_rows": attendance_rows,
            "payments": payments,
            "lessons": lessons,
        },
    )


# UI: Teacher panel
@app.get("/ui/teacher", response_class=HTMLResponse)
def teacher_panel(request: Request, selected_teacher_id: int | None = None, start_date: str | None = None, end_date: str | None = None, db: Session = Depends(get_db)):
    user = request.session.get("user")
    if not user:
        return RedirectResponse(url="/login/teacher", status_code=302)
    if user.get("role") != "teacher":
        # Öğretmen harici biri geldi: kendi paneline yönlendir
        if user.get("role") == "admin":
            return RedirectResponse(url="/dashboard", status_code=302)
        elif user.get("role") == "staff":
            return RedirectResponse(url="/ui/staff", status_code=302)
        else:
            return RedirectResponse(url="/login/teacher", status_code=302)
    current_teacher_id = user.get("teacher_id")
    if not current_teacher_id:
        # Öğretmen ID yoksa hata göster
        return HTMLResponse(content="""
        <!DOCTYPE html>
        <html>
        <head><title>Hata - Piarte</title></head>
        <body>
            <h2>Hata</h2>
            <p>Öğretmen bilgisi bulunamadı. Lütfen yönetici ile iletişime geçin.</p>
            <a href="/logout">Çıkış Yap</a>
        </body>
        </html>
        """, status_code=400)
    try:
        # Ders programında pasif öğrenciler hiç gösterilmez (yoklama/puantaj yine tüm öğrencilerle hesaplanır).
        # Seçilen öğretmen ID'si yoksa, kendi ID'sini kullan
        display_teacher_id = selected_teacher_id if selected_teacher_id else current_teacher_id
        
        # Tarih filtrelerini parse et
        from datetime import date
        start_date_obj = None
        end_date_obj = None
        if start_date:
            try:
                y, m, d = map(int, start_date.split("-"))
                start_date_obj = date(y, m, d)
            except Exception:
                start_date_obj = None
        if end_date:
            try:
                y, m, d = map(int, end_date.split("-"))
                end_date_obj = date(y, m, d)
            except Exception:
                end_date_obj = None
        
        # Tüm öğretmenleri getir
        all_teachers = crud.list_teachers(db)
        
        # Seçilen öğretmenin derslerini getir
        lessons_with_students = crud.lessons_with_students_by_teacher(db, display_teacher_id)
        formatted_lessons = format_lessons_for_schedule(lessons_with_students)
        # Öğretmene atanmış öğrencileri getir
        teacher_students = []
        if current_teacher_id:
            try:
                teacher_students = crud.list_students_by_teacher(db, current_teacher_id, active_only=False)
                # Debug: Eğer öğrenci yoksa, tüm öğrencileri kontrol et
                if not teacher_students:
                    # Tüm öğrencileri getir ve öğretmene atanmış olanları filtrele
                    all_students = crud.list_students(db)
                    for student in all_students:
                        # Öğrencinin bu öğretmene atanıp atanmadığını kontrol et
                        link = db.scalars(
                            select(models.TeacherStudent)
                            .where(
                                models.TeacherStudent.student_id == student.id,
                                models.TeacherStudent.teacher_id == current_teacher_id
                            )
                        ).first()
                        if link:
                            teacher_students.append(student)
            except Exception as e:
                # Hata durumunda boş liste döndür
                import logging
                logging.error(f"Öğrenci listesi hatası: {e}")
                teacher_students = []
        
        # Tüm öğretmenler için haftalık ders programını hazırla (saat bazlı grid için)
        teachers_schedules = build_teachers_schedules(db, all_teachers)
        
        # Puantaj raporunu hesapla (sadece kendi öğretmeni için)
        attendance_report = []
        attendance_totals = None
        if current_teacher_id:
            attendance_report = crud.get_attendance_report_by_teacher(
                db,
                teacher_id=current_teacher_id,
                start_date=start_date_obj,
                end_date=end_date_obj
            )
            # Toplamları hesapla
            if attendance_report and len(attendance_report) > 0:
                teacher_report = attendance_report[0]
                if teacher_report.get("students"):
                    totals = {
                        "total_present": sum(s.get("present", 0) for s in teacher_report["students"]),
                        "total_excused_absent": sum(s.get("excused_absent", 0) for s in teacher_report["students"]),
                        "total_telafi": sum(s.get("telafi", 0) for s in teacher_report["students"]),
                        "total_unexcused_absent": sum(s.get("unexcused_absent", 0) for s in teacher_report["students"]),
                        "total_lessons": sum(s.get("total", 0) for s in teacher_report["students"])
                    }
                    attendance_totals = totals
        
        context = {
            "request": request,
            "lessons_with_students": formatted_lessons,
            "teacher_students": teacher_students,
            "teachers_schedules": teachers_schedules,
            "all_teachers": all_teachers,
            "selected_teacher_id": display_teacher_id,
            "current_teacher_id": current_teacher_id,
            "attendance_report": attendance_report,
            "attendance_totals": attendance_totals,
            "start_date": start_date or "",
            "end_date": end_date or "",
        }
        return templates.TemplateResponse("teacher_panel.html", context)
    except Exception as e:
        import logging
        logging.error(f"Teacher panel error: {e}")
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html>
        <head><title>Hata - Piarte</title></head>
        <body>
            <h2>Öğretmen Paneli Hatası</h2>
            <p>Bir hata oluştu: {str(e)}</p>
            <a href="/logout">Çıkış Yap</a>
        </body>
        </html>
        """, status_code=500)


# UI: Students - create
@app.get("/students/new", response_class=HTMLResponse)
def student_form(request: Request):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    return templates.TemplateResponse("student_new.html", {"request": request})


@app.post("/students/new")
def student_create(
    request: Request,
    first_name: str = Form(...),
    last_name: str = Form(...),
    date_of_birth: str | None = Form(None),
    parent_name: str | None = Form(None),
    parent_phone: str | None = Form(None),
    address: str | None = Form(None),
    phone_primary: str | None = Form(None),
    phone_secondary: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    dob = None
    if date_of_birth:
        try:
            from datetime import date
            y, m, d = map(int, date_of_birth.split("-"))
            dob = date(y, m, d)
        except Exception:
            dob = None
    payload = schemas.StudentCreate(
        first_name=first_name,
        last_name=last_name,
        date_of_birth=dob,
        parent_name=parent_name,
        parent_phone=parent_phone,
        address=address,
        phone_primary=phone_primary,
        phone_secondary=phone_secondary,
    )
    crud.create_student(db, payload)
    return RedirectResponse(url="/dashboard", status_code=302)


@app.post("/students/{student_id}/update")
def student_update(
    student_id: int,
    request: Request,
    first_name: str = Form(...),
    last_name: str = Form(...),
    date_of_birth: str | None = Form(None),
    parent_name: str | None = Form(None),
    parent_phone: str | None = Form(None),
    address: str | None = Form(None),
    phone_primary: str | None = Form(None),
    phone_secondary: str | None = Form(None),
    db: Session = Depends(get_db),
):
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=302)
    dob = None
    if date_of_birth:
        try:
            from datetime import date
            y, m, d = map(int, date_of_birth.split("-"))
            dob = date(y, m, d)
        except Exception:
            dob = None
    payload = schemas.StudentUpdate(
        first_name=first_name,
        last_name=last_name,
        date_of_birth=dob,
        parent_name=parent_name or None,
        parent_phone=parent_phone or None,
        address=address or None,
        phone_primary=phone_primary or None,
        phone_secondary=phone_secondary or None,
    )
    crud.update_student(db, student_id, payload)
    return RedirectResponse(url=f"/ui/students/{student_id}", status_code=status.HTTP_303_SEE_OTHER)


# UI: Teachers - quick create via form
@app.post("/teachers/new")
def teacher_create_form(
    request: Request,
    first_name: str = Form(...),
    last_name: str = Form(...),
    phone: str | None = Form(None),
    email: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    payload = schemas.TeacherCreate(
        first_name=first_name,
        last_name=last_name,
        phone=phone,
        email=email,
    )
    crud.create_teacher(db, payload)
    return RedirectResponse(url="/dashboard", status_code=302)


@app.post("/teachers/{teacher_id}/update")
def teacher_update_form(
    teacher_id: int,
    request: Request,
    first_name: str = Form(...),
    last_name: str = Form(...),
    phone: str | None = Form(None),
    email: str | None = Form(None),
    hourly_rate_try: str | None = Form(None),
    db: Session = Depends(get_db),
):
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=302)
    rate_val = None
    if hourly_rate_try is not None and str(hourly_rate_try).strip() != "":
        try:
            rate_val = float(str(hourly_rate_try).strip().replace(",", "."))
        except ValueError:
            rate_val = None
    payload = schemas.TeacherUpdate(
        first_name=first_name,
        last_name=last_name,
        phone=phone or None,
        email=email or None,
        hourly_rate_try=rate_val,
    )
    crud.update_teacher(db, teacher_id, payload)
    return RedirectResponse(url=f"/ui/teachers/{teacher_id}", status_code=status.HTTP_303_SEE_OTHER)


# UI: Payments - create
@app.get("/payments/new", response_class=HTMLResponse)
def payment_form(
    request: Request,
    db: Session = Depends(get_db),
    student_id: str | None = None,
    return_to: str | None = None,
):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    students = crud.list_students(db, active_only=True)  # Sadece aktif öğrencileri göster
    selected_student_id = None
    if student_id:
        try:
            selected_student_id = int(student_id)
        except (ValueError, TypeError):
            selected_student_id = None
    from datetime import date
    user = request.session.get("user") or {}
    is_staff_user = user.get("role") == "staff"
    resolved_return_to = safe_return_url(return_to, default_panel_url(user))
    return templates.TemplateResponse(
        "payment_new.html",
        {
            "request": request,
            "students": students,
            "selected_student_id": selected_student_id,
            "is_staff_user": is_staff_user,
            "today_iso": date.today().isoformat(),
            "today_display": date.today().strftime("%d.%m.%Y"),
            "return_to": resolved_return_to,
        },
    )


@app.post("/payments/new")
def payment_create(
    request: Request,
    student_id: int = Form(...),
    amount_try: float = Form(...),
    payment_date: str | None = Form(None),
    method: str | None = Form(None),
    note: str | None = Form(None),
    return_to: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    user = request.session.get("user") or {}
    redirect = redirect_teacher(user)
    if redirect:
        return redirect
    from datetime import date
    is_staff_user = user.get("role") == "staff"
    pd = None
    if is_staff_user:
        # Staff kullanıcıları için ödeme tarihi her zaman bugündür.
        pd = date.today()
    elif payment_date:
        try:
            y, m, d = map(int, payment_date.split("-"))
            pd = date(y, m, d)
        except Exception:
            pd = None
    payload = schemas.PaymentCreate(
        student_id=student_id,
        amount_try=amount_try,
        payment_date=pd,
        method=method,
        note=note,
    )
    payment = crud.create_payment(db, payload)
    # Nakit tahsilat → admin mobil bildirimi (başarısız olsa ödeme yine kayıtlı kalır)
    if push_notify and push_notify.is_nakit_method(method):
        student = crud.get_student(db, student_id)
        student_name = (
            f"{student.first_name} {student.last_name}".strip()
            if student
            else f"Öğrenci #{student_id}"
        )
        actor_name = (user.get("full_name") or user.get("username") or "Kullanıcı").strip()
        role = (user.get("role") or "admin").strip()
        if role == "staff":
            actor_label = actor_name
        else:
            actor_label = f"{actor_name} ({role})"
        push_notify.schedule_admin_cash_notify(
            student_name=student_name,
            amount_try=float(amount_try),
            staff_name=actor_label,
            payment_id=getattr(payment, "id", None),
        )
    set_flash_success(request, "Ödeme başarıyla kaydedildi.")
    return RedirectResponse(url=safe_return_url(return_to, default_panel_url(user)), status_code=302)


# UI: Lessons - create and attendance
@app.get("/lessons/new", response_class=HTMLResponse)
def lesson_form(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    courses = crud.list_courses(db)
    teachers = crud.list_teachers(db)
    students = crud.list_students(db, active_only=True)  # Sadece aktif öğrencileri göster
    return templates.TemplateResponse("lesson_new.html", {"request": request, "courses": courses, "teachers": teachers, "students": students})


# UI: Lessons - schedule list
@app.get("/ui/lessons", response_class=HTMLResponse)
def ui_lessons(
    request: Request,
    start: str | None = None,
    end: str | None = None,
    teacher_id: str | None = None,
    course_id: str | None = None,
    student_name: str | None = None,
    show_empty: str | None = None,
    db: Session = Depends(get_db),
):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    from datetime import date
    start_date = None
    end_date = None
    if start:
        try:
            y, m, d = map(int, start.split("-"))
            start_date = date(y, m, d)
        except Exception:
            start_date = None
    if end:
        try:
            y, m, d = map(int, end.split("-"))
            end_date = date(y, m, d)
        except Exception:
            end_date = None

    # Query parametrelerini integer'a çevir (boş string'leri None yap)
    teacher_id_int: int | None = None
    course_id_int: int | None = None
    if teacher_id and str(teacher_id).strip():
        try:
            teacher_id_int = int(str(teacher_id).strip())
        except (ValueError, TypeError):
            teacher_id_int = None
    if course_id and str(course_id).strip():
        try:
            course_id_int = int(str(course_id).strip())
        except (ValueError, TypeError):
            course_id_int = None
    show_empty_lessons = str(show_empty or "").strip().lower() in {"1", "true", "yes", "on"}

    q = db.query(models.Lesson)
    if start_date:
        q = q.filter(models.Lesson.lesson_date >= start_date)
    if end_date:
        q = q.filter(models.Lesson.lesson_date <= end_date)
    if teacher_id_int:
        q = q.filter(models.Lesson.teacher_id == teacher_id_int)
    if course_id_int:
        q = q.filter(models.Lesson.course_id == course_id_int)

    # Öğrenci adına göre filtre: derse kayıtlı öğrenciler üzerinden
    if student_name and student_name.strip():
        term = f"%{student_name.strip()}%"
        from sqlalchemy import or_
        q = (
            q.join(models.LessonStudent, models.LessonStudent.lesson_id == models.Lesson.id)
             .join(models.Student, models.Student.id == models.LessonStudent.student_id)
             .filter(
                 or_(
                     models.Student.first_name.ilike(term),
                     models.Student.last_name.ilike(term),
                     (models.Student.first_name + " " + models.Student.last_name).ilike(term),
                 )
             )
             .distinct()
        )
    elif not show_empty_lessons:
        # Varsayılan: öğrencisi olmayan ders satırlarını listede gizle
        from sqlalchemy import exists
        q = q.filter(exists().where(models.LessonStudent.lesson_id == models.Lesson.id))
    from sqlalchemy.orm import joinedload
    q = q.options(
        joinedload(models.Lesson.lesson_students).joinedload(models.LessonStudent.student),
        joinedload(models.Lesson.course),
        joinedload(models.Lesson.teacher),
    )
    lessons = q.order_by(models.Lesson.lesson_date.asc()).all()
    teachers = crud.list_teachers(db)
    courses = crud.list_courses(db)
    # Ders başına yoklama sayısı (silme uyarısı için)
    attendance_counts = {}
    if lessons:
        from sqlalchemy import func
        lesson_ids = [l.id for l in lessons]
        rows = db.query(models.Attendance.lesson_id, func.count(models.Attendance.id)).filter(
            models.Attendance.lesson_id.in_(lesson_ids)
        ).group_by(models.Attendance.lesson_id).all()
        attendance_counts = {r[0]: r[1] for r in rows}
    return templates.TemplateResponse(
        "lessons_list.html",
        {
            "request": request,
            "lessons": lessons,
            "teachers": teachers,
            "courses": courses,
            "attendance_counts": attendance_counts,
            "start": start or "",
            "end": end or "",
            "teacher_id": teacher_id_int or "",
            "course_id": course_id_int or "",
            "student_name": student_name or "",
            "show_empty": show_empty_lessons,
        },
    )


@app.post("/lessons/new")
def lesson_create(
    request: Request,
    student_id: str = Form(...),
    first_name: str = Form(None),
    last_name: str = Form(None),
    course_id: int = Form(...),
    teacher_id: int = Form(...),
    lesson_date: str = Form(...),
    lesson_weekday: str | None = Form(None),
    start_time: str | None = Form(None),
    end_time: str | None = Form(None),
    description: str | None = Form(None),
    db: Session = Depends(get_db),
):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    # Eğer yeni öğrenci eklenmişse...
    from datetime import date, time as t
    if not student_id:
        if not first_name or not last_name:
            return RedirectResponse(url="/lessons/new", status_code=302)
        student = schemas.StudentCreate(first_name=first_name, last_name=last_name)
        student_db = crud.create_student(db, student)
        student_id = student_db.id
    else:
        student_id = int(student_id)
    # Dersi oluştur
    y, m, d = map(int, lesson_date.split("-"))
    st = None
    et = None
    if start_time:
        hh, mm = map(int, start_time.split(":"))
        st = t(hh, mm)
    if end_time:
        hh, mm = map(int, end_time.split(":"))
        et = t(hh, mm)
    payload = schemas.LessonCreate(
        course_id=course_id,
        teacher_id=teacher_id,
        lesson_date=date(y, m, d),
        start_time=st,
        end_time=et,
        description=description
    )
    lesson = crud.create_lesson(db, payload)
    if lesson_weekday:
        try:
            requested_day = int(lesson_weekday)
            actual_day = lesson.lesson_date.weekday()
            if requested_day != actual_day:
                from datetime import timedelta
                delta = requested_day - actual_day
                lesson.lesson_date = lesson.lesson_date + timedelta(days=delta)
                db.commit()
                db.refresh(lesson)
        except Exception:
            pass
    # Tüm işlemleri tek bir transaction içinde yap
    try:
        # Öğrenciyi, oluşturulan dersin course'una kaydet (commit yapma)
        crud.enroll_student(db, student_id, course_id, commit=False)
        # Öğrenciyi öğretmene ata (eğer atanmamışsa)
        crud.assign_student_to_teacher(db, teacher_id, student_id)
        # Öğrenciyi bu derse özel olarak ata
        crud.assign_student_to_lesson(db, lesson.id, student_id)
        # Tüm değişiklikleri commit et
        db.commit()
    except Exception as e:
        db.rollback()
        # Eğer tablo yoksa, oluştur ve tekrar dene
        try:
            from .db import Base, engine
            Base.metadata.create_all(bind=engine)
            # Tekrar dene
            crud.enroll_student(db, student_id, course_id, commit=False)
            crud.assign_student_to_teacher(db, teacher_id, student_id)
            crud.assign_student_to_lesson(db, lesson.id, student_id)
            db.commit()
        except Exception as e2:
            # Hata mesajını logla
            import logging
            logging.error(f"Ders öğrenci atama hatası: {e2}")
            db.rollback()
            # Hata olsa bile derse yönlendir (ders oluşturuldu)
    
    # Ders oluşturuldu, dashboard'a yönlendir
    return RedirectResponse(url="/dashboard", status_code=302)


@app.get("/lessons/{lesson_id}/attendance/new", response_class=HTMLResponse)
def attendance_form(
    lesson_id: int,
    request: Request,
    db: Session = Depends(get_db),
    error: str | None = None,
    duplicate_warning: str | None = None,
    success: str | None = None,
    attendance_date: str | None = None,
    focus_student: str | None = None,
    return_to: str | None = None,
):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    from datetime import date as date_cls
    lesson = db.get(models.Lesson, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")
    user = request.session.get("user")
    if user.get("role") == "teacher":
        if lesson.teacher_id != user.get("teacher_id"):
            return RedirectResponse(url="/ui/teacher", status_code=302)
        # Pasif öğrenciler de derse bağlıysa listede görünsün (geçmiş yoklama / puantaj tutarlılığı)
        students = crud.list_students_by_lesson(db, lesson_id, active_only=False)
    else:
        students = crud.list_students_by_lesson(db, lesson_id, active_only=False)
    
    # Bu ders için mevcut yoklamaları getir
    existing_attendances = crud.list_attendance_for_lesson(db, lesson_id)
    attendance_map = {att.student_id: att.status for att in existing_attendances}
    
    # Her öğrenci için ödeme durumunu ve mevcut yoklama durumunu kontrol et
    students_with_payment_status = []
    for student in students:
        needs_payment = crud.check_student_payment_status(db, student.id)
        current_status = attendance_map.get(student.id, "")
        students_with_payment_status.append({
            "student": student,
            "needs_payment": needs_payment,
            "current_status": current_status
        })
    
    # Öğretmen için bugünün tarihini, diğerleri için ders tarihini kullan
    if user.get("role") == "teacher":
        default_attendance_date = date_cls.today()
    else:
        default_attendance_date = lesson.lesson_date or date_cls.today()
        if attendance_date and attendance_date.strip():
            try:
                y, m, d = map(int, attendance_date.strip().split("-"))
                default_attendance_date = date_cls(y, m, d)
            except Exception:
                pass

    focus_student_id = None
    if focus_student and focus_student.strip():
        try:
            focus_student_id = int(focus_student.strip())
        except (ValueError, TypeError):
            focus_student_id = None
    
    # Hata mesajını al
    error_message = None
    if error == "no_data" or request.session.get("attendance_errors"):
        error_message = request.session.get("attendance_errors", "Lütfen en az bir öğrenci için durum seçin.")
        request.session.pop("attendance_errors", None)
    
    # Öğretmen için o gün alınan yoklamaları getir
    today_attendances_summary = None
    if user.get("role") == "teacher" and lesson.teacher_id == user.get("teacher_id"):
        from sqlalchemy import func
        from datetime import datetime
        today = date_cls.today()
        today_start = datetime.combine(today, datetime.min.time())
        today_end = datetime.combine(today, datetime.max.time())
        
        # Öğretmenin bugün aldığı tüm yoklamaları getir
        today_attendances = db.scalars(
            select(models.Attendance)
            .join(models.Lesson)
            .where(
                models.Lesson.teacher_id == user.get("teacher_id"),
                models.Attendance.marked_at >= today_start,
                models.Attendance.marked_at <= today_end
            )
            .order_by(models.Attendance.marked_at.desc())
        ).all()
        
        # Özet bilgileri hazırla
        if today_attendances:
            summary_by_lesson = {}
            for att in today_attendances:
                lesson_obj = db.get(models.Lesson, att.lesson_id)
                if not lesson_obj:
                    continue
                
                lesson_key = f"{lesson_obj.id}_{lesson_obj.course.name if lesson_obj.course else 'Bilinmeyen'}"
                if lesson_key not in summary_by_lesson:
                    summary_by_lesson[lesson_key] = {
                        "lesson_id": lesson_obj.id,
                        "course_name": lesson_obj.course.name if lesson_obj.course else "Bilinmeyen",
                        "lesson_time": lesson_obj.start_time.strftime("%H:%M") if lesson_obj.start_time else "",
                        "attendances": [],
                        "counts": {
                            "PRESENT": 0,
                            "EXCUSED_ABSENT": 0,
                            "TELAFI": 0,
                            "UNEXCUSED_ABSENT": 0,
                            "LATE": 0  # Eski kayıtlar için
                        }
                    }
                
                student = db.get(models.Student, att.student_id)
                if student:
                    status = att.status
                    # Eski LATE ve ABSENT değerlerini normalize et
                    if status == "LATE":
                        status = "TELAFI"
                    elif status == "ABSENT":
                        status = "UNEXCUSED_ABSENT"
                    
                    summary_by_lesson[lesson_key]["attendances"].append({
                        "student_name": f"{student.first_name} {student.last_name}",
                        "status": status,
                        "marked_at": att.marked_at.strftime("%H:%M") if att.marked_at else ""
                    })
                    
                    if status in summary_by_lesson[lesson_key]["counts"]:
                        summary_by_lesson[lesson_key]["counts"][status] += 1
            
            today_attendances_summary = list(summary_by_lesson.values())

    resolved_return_to = safe_return_url(return_to, default_panel_url(user))

    return templates.TemplateResponse(
        "attendance_new.html",
        {
            "request": request,
            "lesson": lesson,
            "students_with_status": students_with_payment_status,
            "attendance_date": default_attendance_date.isoformat(),
            "focus_student_id": focus_student_id,
            "error_message": error_message,
            "success_message": success,
            "today_attendances_summary": today_attendances_summary,
            "return_to": resolved_return_to,
        },
    )


@app.get("/lessons/{lesson_id}/attendance/correct", response_class=HTMLResponse)
def correct_attendance_from_schedule(
    lesson_id: int,
    request: Request,
    student_id: int,
    attendance_date: str | None = None,
    return_to: str | None = None,
    db: Session = Depends(get_db),
):
    """Ders programından yoklama düzeltme: öğrencinin tüm yoklamalarını listeler."""
    from datetime import datetime, date as date_cls
    from urllib.parse import urlencode

    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=302)

    lesson = db.get(models.Lesson, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")

    student = db.get(models.Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Öğrenci bulunamadı")

    teacher = db.get(models.Teacher, lesson.teacher_id) if lesson.teacher_id else None
    course = db.get(models.Course, lesson.course_id) if lesson.course_id else None

    attendances = db.scalars(
        select(models.Attendance)
        .where(models.Attendance.student_id == student_id)
        .order_by(models.Attendance.marked_at.desc())
    ).all()

    if not attendances:
        params = {"focus_student": student_id}
        if attendance_date and attendance_date.strip():
            params["attendance_date"] = attendance_date.strip()
        request.session["attendance_errors"] = "Bu öğrenci için yoklama kaydı bulunamadı. Önce yoklama alın."
        return RedirectResponse(
            url=f"/lessons/{lesson_id}/attendance/new?{urlencode(params)}",
            status_code=302,
        )

    highlight_date = None
    if attendance_date and attendance_date.strip():
        try:
            y, m, d = map(int, attendance_date.strip().split("-"))
            highlight_date = date_cls(y, m, d)
        except Exception:
            highlight_date = None

    attendance_rows = []
    for att in attendances:
        att_lesson = db.get(models.Lesson, att.lesson_id)
        att_teacher = db.get(models.Teacher, att_lesson.teacher_id) if att_lesson and att_lesson.teacher_id else None
        att_course = db.get(models.Course, att_lesson.course_id) if att_lesson and att_lesson.course_id else None
        att_date = att.marked_at.date() if att.marked_at else None
        attendance_rows.append({
            "attendance": att,
            "lesson": att_lesson,
            "teacher": att_teacher,
            "course": att_course,
            "highlight": highlight_date is not None and att_date == highlight_date,
        })

    success = request.session.pop("flash_success", None)
    error = request.session.pop("flash_error", None)
    resolved_return_to = safe_return_url(return_to, "/dashboard#ders-programi")

    return templates.TemplateResponse(
        "attendance_correct_list.html",
        {
            "request": request,
            "lesson": lesson,
            "student": student,
            "teacher": teacher,
            "course": course,
            "attendance_rows": attendance_rows,
            "highlight_date": highlight_date,
            "success": success,
            "error": error,
            "return_to": resolved_return_to,
        },
    )


@app.post("/lessons/{lesson_id}/attendance/new")
async def attendance_create(lesson_id: int, request: Request, db: Session = Depends(get_db)):
    import logging
    from datetime import date as date_cls, datetime, time as time_cls

    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    user = request.session.get("user")
    lesson = db.get(models.Lesson, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")
    if user.get("role") == "teacher" and lesson.teacher_id != user.get("teacher_id"):
        return RedirectResponse(url="/ui/teacher", status_code=302)

    lesson_students = crud.list_students_by_lesson(db, lesson_id, active_only=False)
    allowed_student_ids = {s.id for s in lesson_students}
    form = await request.form()
    return_to_value = form.get("return_to")
    default_return_url = default_panel_url(user)
    attendance_date_raw = form.get("attendance_date")
    marked_at_dt = None

    if user.get("role") == "teacher":
        has_telafi = any(
            key.startswith("status_") and (value or "").strip().upper() == "TELAFI"
            for key, value in form.items()
        )
        if not has_telafi:
            today = date_cls.today()
            base_time = lesson.start_time or time_cls(hour=12, minute=0)
            if not isinstance(base_time, time_cls):
                base_time = time_cls(hour=12, minute=0)
            marked_at_dt = datetime.combine(today, base_time)
        elif attendance_date_raw and attendance_date_raw.strip():
            try:
                year, month, day = map(int, attendance_date_raw.split("-"))
                chosen_date = date_cls(year, month, day)
                base_time = lesson.start_time or time_cls(hour=12, minute=0)
                if not isinstance(base_time, time_cls):
                    base_time = time_cls(hour=12, minute=0)
                marked_at_dt = datetime.combine(chosen_date, base_time)
            except Exception:
                today = date_cls.today()
                base_time = lesson.start_time or time_cls(hour=12, minute=0)
                if not isinstance(base_time, time_cls):
                    base_time = time_cls(hour=12, minute=0)
                marked_at_dt = datetime.combine(today, base_time)
        else:
            today = date_cls.today()
            base_time = lesson.start_time or time_cls(hour=12, minute=0)
            if not isinstance(base_time, time_cls):
                base_time = time_cls(hour=12, minute=0)
            marked_at_dt = datetime.combine(today, base_time)
    elif attendance_date_raw:
        try:
            year, month, day = map(int, attendance_date_raw.split("-"))
            chosen_date = date_cls(year, month, day)
            base_time = lesson.start_time or time_cls(hour=12, minute=0)
            if not isinstance(base_time, time_cls):
                base_time = time_cls(hour=12, minute=0)
            marked_at_dt = datetime.combine(chosen_date, base_time)
        except Exception:
            marked_at_dt = None

    valid_statuses = {"PRESENT", "UNEXCUSED_ABSENT", "EXCUSED_ABSENT", "TELAFI"}
    to_create: list[schemas.AttendanceCreate] = []
    passive_attempted_student_names: list[str] = []

    for key, value in form.items():
        if not key.startswith("status_"):
            continue
        try:
            sid = int(key.split("_", 1)[1])
        except Exception:
            continue
        if sid not in allowed_student_ids:
            if user.get("role") == "teacher":
                blocked_student = db.get(models.Student, sid)
                if blocked_student and blocked_student.is_active is False:
                    passive_attempted_student_names.append(
                        f"{blocked_student.first_name} {blocked_student.last_name}"
                    )
            continue

        status_raw = (value or "").strip()
        if not status_raw:
            continue
        status = status_raw.upper()
        if status == "ABSENT":
            status = "UNEXCUSED_ABSENT"
        if status == "LATE":
            status = "TELAFI"
        if status not in valid_statuses:
            continue

        to_create.append(
            schemas.AttendanceCreate(
                lesson_id=lesson_id,
                student_id=sid,
                status=status,
                marked_at=marked_at_dt,
            )
        )

    if marked_at_dt and user.get("role") in ("teacher", "staff") and to_create:
        duplicate_ids = crud.find_attendance_duplicate_student_ids(
            db,
            lesson_id=lesson_id,
            student_ids=[item.student_id for item in to_create],
            attendance_date=marked_at_dt.date(),
        )
        if duplicate_ids:
            duplicate_students = []
            for item in to_create:
                if item.student_id in duplicate_ids:
                    student = db.get(models.Student, item.student_id)
                    if student:
                        duplicate_students.append(f"{student.first_name} {student.last_name}")
            if duplicate_students:
                duplicate_message = (
                    f"Daha önce bu öğrenci{'ler' if len(duplicate_students) > 1 else ''} "
                    f"için yoklama almışsınız: {', '.join(duplicate_students)}"
                )
                request.session["attendance_duplicate_warning"] = duplicate_message
                return RedirectResponse(
                    url=attendance_new_url(lesson_id, return_to_value, duplicate_warning="true"),
                    status_code=302,
                )

    if not to_create:
        if user.get("role") == "teacher" and passive_attempted_student_names:
            names = ", ".join(dict.fromkeys(passive_attempted_student_names))
            request.session["attendance_errors"] = f"Pasif öğrenciler için yoklama alınamaz: {names}"
        elif len(lesson_students) == 0:
            request.session["attendance_errors"] = "Bu derse henüz öğrenci atanmamış. Lütfen önce öğrenci atayın."
        else:
            request.session["attendance_errors"] = (
                "Yoklama verisi bulunamadı. Lütfen en az bir öğrenci için durum seçin "
                "(Geldi, Haberli Gelmedi, Telafi, veya Habersiz Gelmedi)."
            )
        return RedirectResponse(
            url=attendance_new_url(lesson_id, return_to_value, error="no_data"),
            status_code=302,
        )

    try:
        success_count = crud.create_attendances_bulk(db, to_create)
    except Exception as exc:
        db.rollback()
        logging.error("Yoklama kaydedilemedi: %s", exc)
        request.session["attendance_errors"] = "Yoklama kaydedilirken bir hata oluştu."
        return RedirectResponse(
            url=attendance_new_url(lesson_id, return_to_value, error="no_data"),
            status_code=302,
        )

    if success_count == 1:
        success_msg = "Yoklama başarıyla kaydedildi."
    else:
        success_msg = f"{success_count} öğrenci için yoklama kaydedildi."
    set_flash_success(request, success_msg)
    return RedirectResponse(
        url=safe_return_url(return_to_value, default_return_url),
        status_code=302,
    )


@app.post("/attendances/{attendance_id}/delete")
def delete_attendance_endpoint(
	attendance_id: int,
	request: Request,
	db: Session = Depends(get_db),
):
	"""Tek bir yoklama kaydını sil (sadece admin)"""
	if not request.session.get("user"):
		return RedirectResponse(url="/", status_code=302)
	user = request.session.get("user")
	if user.get("role") != "admin":
		raise HTTPException(status_code=403, detail="Sadece admin bu işlemi yapabilir")

	return_to = request.query_params.get("return_to")

	try:
		attendance = crud.delete_attendance(db, attendance_id)
		if attendance:
			import logging
			logging.warning(f"Yoklama kaydı silindi: ID={attendance_id}, Öğrenci={attendance.student_id}, Ders={attendance.lesson_id}")
			msg = "Yoklama kaydı başarıyla silindi"
			if return_to:
				set_flash_success(request, msg)
			else:
				request.session["delete_attendance_success"] = msg
		else:
			request.session["delete_attendance_error"] = "Yoklama kaydı bulunamadı"
	except Exception as e:
		import logging
		import traceback
		logging.error(f"Yoklama kaydı silinirken hata: {e}")
		logging.error(traceback.format_exc())
		request.session["delete_attendance_error"] = str(e)
	
	# Filtreleri koruyarak dashboard'a yönlendir
	from urllib.parse import urlencode
	if return_to:
		redirect_url = safe_return_url(return_to, "/dashboard")
	else:
		params = {}
		if request.query_params.get("teacher_id"):
			params["teacher_id"] = request.query_params.get("teacher_id")
		if request.query_params.get("student_id"):
			params["student_id"] = request.query_params.get("student_id")
		if request.query_params.get("course_id"):
			params["course_id"] = request.query_params.get("course_id")
		if request.query_params.get("status"):
			params["status"] = request.query_params.get("status")
		if request.query_params.get("start_date"):
			params["start_date"] = request.query_params.get("start_date")
		if request.query_params.get("end_date"):
			params["end_date"] = request.query_params.get("end_date")
		if request.query_params.get("order_by"):
			params["order_by"] = request.query_params.get("order_by")
		redirect_url = "/dashboard"
		if params:
			redirect_url += "?" + urlencode(params)

	return RedirectResponse(url=redirect_url, status_code=302)


@app.get("/attendances/{attendance_id}/edit", response_class=HTMLResponse)
def edit_attendance_form(
	attendance_id: int,
	request: Request,
	from_lesson: int | None = None,
	from_student: int | None = None,
	return_to: str | None = None,
	db: Session = Depends(get_db),
):
	"""Yoklama düzenleme formu (staff için)"""
	user = request.session.get("user")
	if not user or user.get("role") not in ["admin", "staff"]:
		return RedirectResponse(url="/", status_code=302)
	
	attendance = db.get(models.Attendance, attendance_id)
	if not attendance:
		request.session["error"] = "Yoklama kaydı bulunamadı"
		if user.get("role") == "admin":
			return RedirectResponse(url="/dashboard", status_code=302)
		return RedirectResponse(url="/ui/staff", status_code=302)
	
	lesson = db.get(models.Lesson, attendance.lesson_id)
	student = db.get(models.Student, attendance.student_id)
	teacher = db.get(models.Teacher, lesson.teacher_id) if lesson and lesson.teacher_id else None
	course = db.get(models.Course, lesson.course_id) if lesson and lesson.course_id else None
	courses = crud.list_courses(db)
	resolved_return_to = safe_return_url(return_to, default_panel_url(user))

	return templates.TemplateResponse("attendance_edit.html", {
		"request": request,
		"attendance": attendance,
		"lesson": lesson,
		"student": student,
		"teacher": teacher,
		"course": course,
		"courses": courses,
		"from_lesson_id": from_lesson,
		"from_student_id": from_student,
		"return_to": resolved_return_to,
	})


@app.post("/attendances/{attendance_id}/edit")
def update_attendance_endpoint(
	attendance_id: int,
	request: Request,
	status: str = Form(...),
	marked_at_date: str = Form(...),
	marked_at_time: str | None = Form(None),
	note: str | None = Form(None),
	course_id: str | None = Form(None),
	from_lesson_id: str | None = Form(None),
	from_student_id: str | None = Form(None),
	return_to: str | None = Form(None),
	db: Session = Depends(get_db),
):
	"""Yoklama kaydını güncelle (staff için). Kurs değişirse sadece bu kayıt etkilenir."""
	user = request.session.get("user")
	if not user or user.get("role") not in ["admin", "staff"]:
		return RedirectResponse(url="/", status_code=302)
	
	try:
		from datetime import datetime
		from sqlalchemy import select, func
		
		attendance = db.get(models.Attendance, attendance_id)
		if not attendance:
			request.session["error"] = "Yoklama kaydı bulunamadı"
			if user.get("role") == "admin":
				return RedirectResponse(url="/dashboard", status_code=302)
			return RedirectResponse(url="/ui/staff", status_code=302)
		
		lesson = db.get(models.Lesson, attendance.lesson_id)
		new_course_id = int(course_id) if course_id and course_id.strip() and course_id.isdigit() else None
		
		# Kurs değiştiyse: sadece bu yoklamayı etkile (başka öğrencilere dokunma)
		if new_course_id and lesson and lesson.course_id != new_course_id:
			count = db.scalar(select(func.count(models.Attendance.id)).where(models.Attendance.lesson_id == lesson.id))
			if count == 1:
				# Ders sadece bu yoklamaya ait; dersin kursunu güncelle
				crud.update_lesson(db, lesson.id, schemas.LessonUpdate(course_id=new_course_id))
			else:
				# Başka yoklamalar da var; yeni ders oluştur (aynı tarih/öğretmen/saat), bu yoklamayı taşı
				new_lesson = crud.create_lesson(db, schemas.LessonCreate(
					course_id=new_course_id,
					teacher_id=lesson.teacher_id,
					lesson_date=lesson.lesson_date,
					start_time=lesson.start_time,
					end_time=lesson.end_time,
					description=lesson.description,
				))
				attendance.lesson_id = new_lesson.id
				db.commit()
				db.refresh(attendance)
		
		# Tarih ve saat bilgisini birleştir
		marked_at_datetime = None
		if marked_at_time:
			try:
				hour, minute = map(int, marked_at_time.split(":"))
				marked_at_datetime = datetime.combine(
					datetime.strptime(marked_at_date, "%Y-%m-%d").date(),
					datetime.min.time().replace(hour=hour, minute=minute)
				)
			except (ValueError, AttributeError):
				marked_at_datetime = datetime.combine(
					datetime.strptime(marked_at_date, "%Y-%m-%d").date(),
					datetime.min.time()
				)
		else:
			marked_at_datetime = datetime.combine(
				datetime.strptime(marked_at_date, "%Y-%m-%d").date(),
				datetime.min.time()
			)
		
		# Yoklama kaydını güncelle (durum, tarih, not)
		updated_attendance = crud.update_attendance(
			db,
			attendance_id=attendance_id,
			status=status,
			marked_at=marked_at_datetime,
			note=note
		)
		
		if updated_attendance:
			set_flash_success(request, "Yoklama kaydı başarıyla güncellendi.")
		else:
			request.session["flash_error"] = "Yoklama kaydı bulunamadı"
	except Exception as e:
		import logging
		import traceback
		logging.error(f"Yoklama güncellenirken hata: {e}")
		logging.error(traceback.format_exc())
		request.session["flash_error"] = f"Yoklama güncellenirken hata oluştu: {str(e)}"

	if from_lesson_id and from_lesson_id.strip().isdigit() and from_student_id and from_student_id.strip().isdigit():
		return RedirectResponse(
			url=f"/lessons/{from_lesson_id.strip()}/attendance/correct?student_id={from_student_id.strip()}",
			status_code=302,
		)
	return RedirectResponse(
		url=safe_return_url(return_to, default_panel_url(user)),
		status_code=302,
	)


@app.post("/admin/clear-all-attendances")
def clear_all_attendances(request: Request, db: Session = Depends(get_db)):
	"""Tüm yoklama kayıtlarını sil (sadece admin)"""
	if not request.session.get("user"):
		return RedirectResponse(url="/", status_code=302)
	user = request.session.get("user")
	if user.get("role") != "admin":
		raise HTTPException(status_code=403, detail="Sadece admin bu işlemi yapabilir")
	
	try:
		count = crud.delete_all_attendances(db)
		import logging
		logging.warning(f"Tüm yoklama kayıtları silindi: {count} kayıt")
		request.session["clear_attendances_success"] = f"{count} yoklama kaydı silindi"
		return RedirectResponse(url="/dashboard", status_code=302)
	except Exception as e:
		import logging
		import traceback
		logging.error(f"Yoklama kayıtları silinirken hata: {e}")
		logging.error(traceback.format_exc())
		request.session["clear_attendances_error"] = str(e)
		return RedirectResponse(url="/dashboard", status_code=302)


# UI: Enrollment - create
@app.get("/enrollments/new", response_class=HTMLResponse)
def enrollment_form(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    students = crud.list_students(db, active_only=True)  # Sadece aktif öğrencileri göster
    courses = crud.list_courses(db)
    return templates.TemplateResponse("enrollment_new.html", {"request": request, "students": students, "courses": courses})


@app.post("/enrollments/new")
def enrollment_create(request: Request, student_id: int = Form(...), course_id: int = Form(...), db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    try:
        crud.enroll_student(db, student_id, course_id)
    except Exception:
        pass
    return RedirectResponse(url="/dashboard", status_code=302)


# Students
@app.post("/students", response_model=schemas.StudentOut)
def create_student(payload: schemas.StudentCreate, db: Session = Depends(get_db)):
	return crud.create_student(db, payload)


@app.get("/students", response_model=list[schemas.StudentOut])
def list_students(db: Session = Depends(get_db)):
	return crud.list_students(db)


@app.get("/api/students/search")
def search_students(q: str = None, db: Session = Depends(get_db)):
	"""Öğrenci arama API endpoint'i - autocomplete için (en az 3 harf)"""
	if not q or len(q.strip()) < 3:
		return []
	from sqlalchemy import or_
	search_prefix = f"{q.strip()[:3]}%"
	students = db.query(models.Student).filter(
		or_(
			models.Student.first_name.ilike(search_prefix),
			models.Student.last_name.ilike(search_prefix),
			(models.Student.first_name + " " + models.Student.last_name).ilike(search_prefix),
		)
	).limit(10).all()
	return [
		{
			"id": s.id,
			"first_name": s.first_name,
			"last_name": s.last_name,
			"full_name": f"{s.first_name} {s.last_name}",
			"phone": s.phone_primary or s.phone_secondary or None,
			"type": "student"
		}
		for s in students
	]


@app.get("/api/teachers/search")
def search_teachers(q: str = None, db: Session = Depends(get_db)):
	"""Öğretmen arama API endpoint'i - autocomplete için"""
	if not q or len(q.strip()) < 3:
		return []
	search_prefix = f"{q.strip()[:3]}%"
	teachers = db.query(models.Teacher).filter(
		models.Teacher.is_active == True,
		(models.Teacher.first_name.ilike(search_prefix)) | 
		(models.Teacher.last_name.ilike(search_prefix))
	).limit(10).all()
	return [
		{
			"id": t.id,
			"first_name": t.first_name,
			"last_name": t.last_name,
			"full_name": f"{t.first_name} {t.last_name}",
			"type": "teacher"
		}
		for t in teachers
	]


@app.get("/api/courses/search")
def search_courses(q: str = None, db: Session = Depends(get_db)):
	"""Kurs arama API endpoint'i - autocomplete için"""
	if not q or len(q.strip()) < 3:
		return []
	search_term = f"%{q.strip()}%"
	courses = db.query(models.Course).filter(
		models.Course.name.ilike(search_term)
	).limit(10).all()
	return [
		{
			"id": c.id,
			"name": c.name,
			"type": "course"
		}
		for c in courses
	]


@app.get("/api/search/all")
def search_all(q: str = None, db: Session = Depends(get_db)):
	"""İsim bazlı arama API endpoint'i - autocomplete için (öğrenci, öğretmen)"""
	if not q or len(q.strip()) < 3:
		return []
	search_prefix = f"{q.strip()[:3]}%"
	results = []
	
	# Öğrenciler
	students = db.query(models.Student).filter(
		(models.Student.first_name.ilike(search_prefix)) | 
		(models.Student.last_name.ilike(search_prefix))
	).limit(5).all()
	for s in students:
		results.append({
			"id": s.id,
			"name": f"{s.first_name} {s.last_name}",
			"type": "student",
			"url": f"/ui/search/student/{s.id}?return_to=/dashboard"
		})
	
	# Öğretmenler
	teachers = db.query(models.Teacher).filter(
		models.Teacher.is_active == True,
		(models.Teacher.first_name.ilike(search_prefix)) |
		(models.Teacher.last_name.ilike(search_prefix))
	).limit(5).all()
	for t in teachers:
		results.append({
			"id": t.id,
			"name": f"{t.first_name} {t.last_name}",
			"type": "teacher",
			"url": f"/ui/teachers/{t.id}"
		})
	
	return results


@app.patch("/students/{student_id}", response_model=schemas.StudentOut)
def update_student(student_id: int, payload: schemas.StudentUpdate, db: Session = Depends(get_db)):
	student = crud.update_student(db, student_id, payload)
	if not student:
		raise HTTPException(status_code=404, detail="Öğrenci bulunamadı")
	return student


# Teachers
@app.post("/teachers", response_model=schemas.TeacherOut)
def create_teacher(payload: schemas.TeacherCreate, db: Session = Depends(get_db)):
	return crud.create_teacher(db, payload)


@app.get("/teachers", response_model=list[schemas.TeacherOut])
def list_teachers(db: Session = Depends(get_db)):
	return crud.list_teachers(db)


# Courses
@app.get("/courses", response_model=list[schemas.CourseOut])
def list_courses(db: Session = Depends(get_db)):
	return crud.list_courses(db)


# UI: Courses - list
@app.get("/ui/courses", response_class=HTMLResponse)
def ui_courses(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    if request.session.get("user").get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    courses = crud.list_courses(db)
    return templates.TemplateResponse("courses_list.html", {"request": request, "courses": courses})


# UI: Courses - create form
@app.get("/courses/new", response_class=HTMLResponse)
def course_form(request: Request):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    return templates.TemplateResponse("course_new.html", {"request": request})


# UI: Courses - create
@app.post("/courses/new")
def course_create(
    request: Request,
    name: str = Form(...),
    db: Session = Depends(get_db),
):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    payload = schemas.CourseCreate(name=name)
    try:
        crud.create_course_from_schema(db, payload)
    except Exception:
        # Kurs adı zaten varsa hata ver
        pass
    return RedirectResponse(url="/dashboard", status_code=302)


# UI: Courses - update form
@app.get("/courses/{course_id}/edit", response_class=HTMLResponse)
def course_edit_form(course_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    redirect = redirect_teacher(request.session.get("user"))
    if redirect:
        return redirect
    course = crud.get_course(db, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Kurs bulunamadı")
    return templates.TemplateResponse("course_edit.html", {"request": request, "course": course})


# UI: Courses - update
@app.post("/courses/{course_id}/update")
def course_update(
    course_id: int,
    request: Request,
    name: str = Form(...),
    db: Session = Depends(get_db),
):
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=302)
    payload = schemas.CourseUpdate(name=name)
    crud.update_course(db, course_id, payload)
    return RedirectResponse(url="/dashboard", status_code=status.HTTP_303_SEE_OTHER)


# UI: Courses - delete
@app.post("/courses/{course_id}/delete")
def course_delete(course_id: int, request: Request, db: Session = Depends(get_db)):
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=status.HTTP_303_SEE_OTHER)
    crud.delete_course(db, course_id)
    return RedirectResponse(url="/dashboard", status_code=status.HTTP_303_SEE_OTHER)


# Enrollment
@app.post("/enrollments")
def enroll_student(payload: schemas.EnrollmentCreate, db: Session = Depends(get_db)):
	return crud.enroll_student(db, payload.student_id, payload.course_id)


# Lessons
@app.post("/lessons", response_model=schemas.LessonOut)
def create_lesson(payload: schemas.LessonCreate, db: Session = Depends(get_db)):
	return crud.create_lesson(db, payload)


@app.get("/teachers/{teacher_id}/lessons", response_model=list[schemas.LessonOut])
def lessons_by_teacher(teacher_id: int, db: Session = Depends(get_db)):
	return crud.list_lessons_by_teacher(db, teacher_id)


# UI: Lessons - edit form
@app.get("/lessons/{lesson_id}/edit", response_class=HTMLResponse)
def lesson_edit_form(lesson_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    user = request.session.get("user")
    # Ders programı düzenleme: admin ve staff yetkili
    if user.get("role") not in ["admin", "staff"]:
        return RedirectResponse(url="/login/admin", status_code=302)
    lesson = crud.get_lesson(db, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")
    courses = crud.list_courses(db)
    teachers = crud.list_teachers(db)
    return templates.TemplateResponse("lesson_edit.html", {"request": request, "lesson": lesson, "courses": courses, "teachers": teachers})


# UI: Lessons - update
@app.post("/lessons/{lesson_id}/update")
def lesson_update(
    lesson_id: int,
    request: Request,
    course_id: int = Form(...),
    teacher_id: int = Form(...),
    lesson_date: str = Form(...),
    start_time: str | None = Form(None),
    end_time: str | None = Form(None),
    description: str | None = Form(None),
    db: Session = Depends(get_db),
):
    user = request.session.get("user")
    # Ders programı düzenleme: admin ve staff yetkili
    if not user or user.get("role") not in ["admin", "staff"]:
        return RedirectResponse(url="/login/admin", status_code=302)
    from datetime import date, time as t
    y, m, d = map(int, lesson_date.split("-"))
    st = None
    et = None
    if start_time:
        try:
            hh, mm = map(int, start_time.split(":"))
            st = t(hh, mm)
        except Exception:
            st = None
    if end_time:
        try:
            hh, mm = map(int, end_time.split(":"))
            et = t(hh, mm)
        except Exception:
            et = None
    payload = schemas.LessonUpdate(
        course_id=course_id,
        teacher_id=teacher_id,
        lesson_date=date(y, m, d),
        start_time=st,
        end_time=et,
        description=description
    )
    lesson = crud.update_lesson(db, lesson_id, payload)
    if not lesson:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")
    # Öğretmen detay sayfasına yönlendir
    return RedirectResponse(url=f"/ui/teachers/{teacher_id}", status_code=status.HTTP_303_SEE_OTHER)


# UI: Lessons - add student form
@app.get("/lessons/{lesson_id}/add-student", response_class=HTMLResponse)
def lesson_add_student_form(lesson_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    user = request.session.get("user")
    if user.get("role") not in ["admin", "staff"]:
        return RedirectResponse(url="/login/admin", status_code=302)
    lesson = crud.get_lesson(db, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")
    # Derse atanmamış öğrencileri getir (course'a kayıtlı ama derse atanmamış) - sadece aktif öğrenciler
    enrolled_students = db.scalars(
        select(models.Student)
        .join(models.Enrollment, models.Enrollment.student_id == models.Student.id)
        .where(
            models.Enrollment.course_id == lesson.course_id,
            models.Student.is_active == True
        )
    ).all()
    assigned_student_ids = {s.id for s in crud.list_students_by_lesson(db, lesson_id)}
    available_students = [s for s in enrolled_students if s.id not in assigned_student_ids]
    # Tüm aktif öğrencileri de seçenek olarak ekle
    all_students = crud.list_students(db, active_only=True)
    return templates.TemplateResponse("lesson_add_student.html", {
        "request": request,
        "lesson": lesson,
        "available_students": available_students,
        "all_students": all_students
    })


# UI: Lessons - add student
@app.post("/lessons/{lesson_id}/add-student")
def lesson_add_student(lesson_id: int, request: Request, student_id: int = Form(...), db: Session = Depends(get_db)):
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "staff"]:
        return RedirectResponse(url="/login/admin", status_code=302)
    lesson = crud.get_lesson(db, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")
    # Öğrenciyi derse ata
    crud.assign_student_to_lesson(db, lesson_id, student_id)
    # Öğrenciyi course'a kaydet (eğer kayıtlı değilse)
    try:
        crud.enroll_student(db, student_id, lesson.course_id)
    except Exception:
        pass  # Zaten kayıtlı olabilir
    # Öğrenciyi öğretmene ata (eğer atanmamışsa)
    crud.assign_student_to_teacher(db, lesson.teacher_id, student_id)
    db.commit()
    return RedirectResponse(url=f"/ui/teachers/{lesson.teacher_id}", status_code=status.HTTP_303_SEE_OTHER)


# UI: Lessons - remove student from slot (attendance history preserved)
@app.post("/lessons/{lesson_id}/remove-student")
def lesson_remove_student(
    lesson_id: int,
    request: Request,
    student_id: int = Form(...),
    return_to: str | None = Form(None),
    db: Session = Depends(get_db),
):
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "staff"]:
        return RedirectResponse(url="/login/admin", status_code=302)
    lesson = crud.get_lesson(db, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")
    removed = crud.remove_student_from_lesson(db, lesson_id, student_id)
    if removed:
        db.commit()
        set_flash_success(request, "Öğrenci dersten çıkarıldı. Geçmiş yoklama kayıtları korundu.")
    default_url = f"/ui/teachers/{lesson.teacher_id}" if lesson.teacher_id else "/ui/lessons"
    return RedirectResponse(
        url=safe_return_url(return_to, default_url),
        status_code=status.HTTP_303_SEE_OTHER,
    )


# UI: Lessons - delete
@app.post("/lessons/{lesson_id}/delete")
def lesson_delete(lesson_id: int, request: Request, db: Session = Depends(get_db)):
    user = request.session.get("user")
    # Ders programı düzenleme: admin ve staff yetkili
    if not user or user.get("role") not in ["admin", "staff"]:
        return RedirectResponse(url="/login/admin", status_code=status.HTTP_303_SEE_OTHER)
    lesson = crud.get_lesson(db, lesson_id)
    if not lesson:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")
    result = crud.delete_lesson(db, lesson_id)
    if result is False:
        # Bu derse ait yoklama kaydı var; silme engellendi
        request.session["lesson_delete_error"] = "Bu derse ait yoklama kayıtları bulunduğu için ders silinemiyor. Yoklamaları korumak için önce ilgili yoklamaları panelden silin."
        return RedirectResponse(url="/ui/lessons", status_code=status.HTTP_303_SEE_OTHER)
    # Program düzenleme ekranına geri dön
    return RedirectResponse(url="/ui/lessons", status_code=status.HTTP_303_SEE_OTHER)


# Attendance
@app.post("/attendance", response_model=schemas.AttendanceOut)
def mark_attendance(payload: schemas.AttendanceCreate, db: Session = Depends(get_db)):
	return crud.mark_attendance(db, payload)


@app.get("/lessons/{lesson_id}/attendance", response_model=list[schemas.AttendanceOut])
def attendance_for_lesson(lesson_id: int, db: Session = Depends(get_db)):
	return crud.list_attendance_for_lesson(db, lesson_id)


# Payments
@app.post("/payments", response_model=schemas.PaymentOut)
def create_payment(payload: schemas.PaymentCreate, db: Session = Depends(get_db)):
	return crud.create_payment(db, payload)


@app.get("/students/{student_id}/payments", response_model=list[schemas.PaymentOut])
def payments_by_student(student_id: int, db: Session = Depends(get_db)):
	return crud.list_payments_by_student(db, student_id)


# UI: Students - list and detail
@app.get("/ui/students", response_class=HTMLResponse)
def ui_students(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    if request.session.get("user").get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    students = crud.list_students(db)
    return templates.TemplateResponse("students_list.html", {"request": request, "students": students})


@app.get("/ui/students/{student_id}", response_class=HTMLResponse)
def ui_student_detail(student_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    if request.session.get("user").get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    student = crud.get_student(db, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="Öğrenci bulunamadı")
    payments = crud.list_payments_by_student(db, student_id)
    # enrollments and courses
    enrollments = db.query(models.Enrollment).filter(models.Enrollment.student_id == student_id).all()
    
    # Öğrencinin yoklama kayıtlarını getir (detaylı bilgilerle)
    attendances_raw = crud.list_all_attendances(db, student_id=student_id, limit=1000, order_by="marked_at_desc")
    
    # Yoklama kayıtlarını detaylı bilgilerle formatla
    attendances_with_details = []
    for att in attendances_raw:
        lesson = db.get(models.Lesson, att.lesson_id) if att.lesson_id else None
        course = None
        teacher = None
        if lesson:
            course = db.get(models.Course, lesson.course_id) if lesson.course_id else None
            teacher = db.get(models.Teacher, lesson.teacher_id) if lesson.teacher_id else None
        
        attendances_with_details.append({
            "attendance": att,
            "lesson": lesson,
            "course": course,
            "teacher": teacher,
        })
    
    return templates.TemplateResponse("student_detail.html", {
        "request": request,
        "student": student,
        "payments": payments,
        "enrollments": enrollments,
        "attendances": attendances_with_details
    })


# UI: Teachers - list and detail
@app.get("/ui/teachers", response_class=HTMLResponse)
def ui_teachers(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    if request.session.get("user").get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    teachers = crud.list_teachers(db)
    return templates.TemplateResponse("teachers_list.html", {"request": request, "teachers": teachers})


@app.get("/ui/teachers/{teacher_id}", response_class=HTMLResponse)
def ui_teacher_detail(teacher_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    if request.session.get("user").get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    teacher = db.get(models.Teacher, teacher_id)
    if not teacher:
        raise HTTPException(status_code=404, detail="Öğretmen bulunamadı")
    lessons = crud.list_lessons_by_teacher(db, teacher_id)
    # Her ders için öğrencileri ve yoklama sayısını getir
    lessons_with_students = []
    for lesson in lessons:
        students = crud.list_students_by_lesson(db, lesson.id, active_only=False)
        students = filter_students_by_passive_flag(students, False)
        if not students:
            # Öğrencisi olmayan dersleri detay program tablosunda gizle
            continue
        att_count = len(crud.list_attendance_for_lesson(db, lesson.id))
        lessons_with_students.append({"lesson": lesson, "students": students, "attendance_count": att_count})
    teacher_students = crud.list_students_by_teacher(db, teacher_id, active_only=False)
    return templates.TemplateResponse("teacher_detail.html", {"request": request, "teacher": teacher, "lessons_with_students": lessons_with_students, "teacher_students": teacher_students})


def _parse_optional_date(value: str | None):
    from datetime import date as date_cls
    if not value or not str(value).strip():
        return None
    try:
        y, m, d = map(int, str(value).strip().split("-"))
        return date_cls(y, m, d)
    except Exception:
        return None


def _default_finance_range(start: str | None, end: str | None):
    """Tarih boşsa içinde bulunulan ay."""
    from datetime import date as date_cls
    from calendar import monthrange
    today = date_cls.today()
    start_date = _parse_optional_date(start)
    end_date = _parse_optional_date(end)
    if start_date is None and end_date is None:
        start_date = today.replace(day=1)
        end_date = today.replace(day=monthrange(today.year, today.month)[1])
    return start_date, end_date, (start_date.isoformat() if start_date else ""), (end_date.isoformat() if end_date else "")


# UI: Finans (admin only)
@app.get("/ui/finance", response_class=HTMLResponse)
def ui_finance(request: Request, start: str | None = None, end: str | None = None, db: Session = Depends(get_db)):
    require_admin(request)
    start_date, end_date, start_s, end_s = _default_finance_range(start, end)
    income_by_method = crud.sum_payments_by_method(db, start_date=start_date, end_date=end_date)
    income_total = crud.sum_payments_total(db, start_date=start_date, end_date=end_date)
    expense_total = crud.sum_expenses(db, start_date=start_date, end_date=end_date)
    net = income_total - expense_total
    income_monthly = crud.monthly_payment_totals(db, start_date=start_date, end_date=end_date)
    expense_monthly = crud.monthly_expense_totals(db, start_date=start_date, end_date=end_date)
    expense_by_category = crud.expense_totals_by_category(db, start_date=start_date, end_date=end_date)
    by_teacher = crud.payment_totals_by_teacher(db, start_date=start_date, end_date=end_date)
    return templates.TemplateResponse(
        "finance.html",
        {
            "request": request,
            "start": start_s,
            "end": end_s,
            "income_total": income_total,
            "expense_total": expense_total,
            "net": net,
            "nakit": income_by_method.get("Nakit", 0),
            "iban": income_by_method.get("EFT", 0),
            "kart": income_by_method.get("Kart", 0),
            "income_monthly": income_monthly,
            "expense_monthly": expense_monthly,
            "expense_by_category": expense_by_category,
            "by_teacher": by_teacher,
        },
    )


@app.get("/ui/finance/income", response_class=HTMLResponse)
def ui_finance_income(
    request: Request,
    start: str | None = None,
    end: str | None = None,
    method: str | None = None,
    teacher_id: str | None = None,
    db: Session = Depends(get_db),
):
    require_admin(request)
    start_date, end_date, start_s, end_s = _default_finance_range(start, end)
    method_filter = (method or "").strip() or None
    # UI "IBAN" -> DB "EFT"
    db_method = None
    if method_filter == "IBAN":
        db_method = "EFT"
    elif method_filter in ("Nakit", "Kart", "EFT"):
        db_method = method_filter

    teacher_id_int = None
    if teacher_id and str(teacher_id).strip():
        try:
            teacher_id_int = int(str(teacher_id).strip())
        except (ValueError, TypeError):
            teacher_id_int = None

    q = db.query(models.Payment).join(models.Student)
    if start_date:
        q = q.filter(models.Payment.payment_date >= start_date)
    if end_date:
        q = q.filter(models.Payment.payment_date <= end_date)
    if db_method:
        q = q.filter(models.Payment.method == db_method)
    if teacher_id_int is not None:
        q = q.join(
            models.TeacherStudent,
            models.TeacherStudent.student_id == models.Payment.student_id,
        ).filter(models.TeacherStudent.teacher_id == teacher_id_int)
    items = q.order_by(models.Payment.payment_date.desc()).all()

    teacher_names = crud.student_teacher_name_map(db)
    payment_rows = [
        {
            "payment": p,
            "teacher_name": teacher_names.get(p.student_id) or "Atanmamış",
        }
        for p in items
    ]

    income_by_method = crud.sum_payments_by_method(db, start_date=start_date, end_date=end_date)
    income_total = crud.sum_payments_total(db, start_date=start_date, end_date=end_date)
    filtered_total = float(sum(float(p.amount_try or 0) for p in items))
    income_monthly = crud.monthly_payment_totals(db, start_date=start_date, end_date=end_date)
    by_teacher = crud.payment_totals_by_teacher(
        db, start_date=start_date, end_date=end_date, method=db_method
    )
    teachers = crud.list_teachers(db)

    return templates.TemplateResponse(
        "finance_income.html",
        {
            "request": request,
            "start": start_s,
            "end": end_s,
            "method": method_filter or "",
            "teacher_id": teacher_id_int or "",
            "teachers": teachers,
            "payment_rows": payment_rows,
            "income_total": income_total,
            "filtered_total": filtered_total,
            "nakit": income_by_method.get("Nakit", 0),
            "iban": income_by_method.get("EFT", 0),
            "kart": income_by_method.get("Kart", 0),
            "income_monthly": income_monthly,
            "by_teacher": by_teacher,
        },
    )


@app.get("/ui/finance/payment-detail", response_class=HTMLResponse)
def ui_finance_payment_detail(
    request: Request,
    start: str | None = None,
    end: str | None = None,
    coverage_start: str | None = None,
    coverage_end: str | None = None,
    student_id: str | None = None,
    teacher_id: str | None = None,
    only_cross_month: str | None = None,
    db: Session = Depends(get_db),
):
    require_admin(request)
    start_date, end_date, start_s, end_s = _default_finance_range(start, end)
    cov_start = _parse_optional_date(coverage_start)
    cov_end = _parse_optional_date(coverage_end)

    student_id_int = None
    if student_id and str(student_id).strip():
        try:
            student_id_int = int(str(student_id).strip())
        except (ValueError, TypeError):
            student_id_int = None
    teacher_id_int = None
    if teacher_id and str(teacher_id).strip():
        try:
            teacher_id_int = int(str(teacher_id).strip())
        except (ValueError, TypeError):
            teacher_id_int = None

    analysis = crud.build_payment_package_details(
        db,
        payment_start=start_date,
        payment_end=end_date,
        coverage_start=cov_start,
        coverage_end=cov_end,
        student_id=student_id_int,
        teacher_id=teacher_id_int,
    )
    rows = analysis["rows"]
    if (only_cross_month or "").strip() in {"1", "true", "on", "yes"}:
        rows = [r for r in rows if r.get("crosses_month")]
        analysis = dict(analysis)
        analysis["rows"] = rows
        # Yeniden özet
        view_cash: dict[str, float] = {}
        view_accrual: dict[str, float] = {}
        view_prepaid: dict[str, float] = {}
        for row in rows:
            pm = row["payment_month"]
            view_cash[pm] = view_cash.get(pm, 0.0) + row["amount"]
            for ym, val in row["month_split"].items():
                if "(bekleyen)" in ym:
                    base = ym.replace(" (bekleyen)", "")
                    view_prepaid[base] = view_prepaid.get(base, 0.0) + val
                else:
                    view_accrual[ym] = view_accrual.get(ym, 0.0) + val
        months = sorted(set(view_cash) | set(view_accrual) | set(view_prepaid))
        analysis["monthly_compare"] = [
            {
                "month": m,
                "cash": round(view_cash.get(m, 0.0), 2),
                "accrued": round(view_accrual.get(m, 0.0), 2),
                "prepaid": round(view_prepaid.get(m, 0.0), 2),
                "delta": round(view_cash.get(m, 0.0) - view_accrual.get(m, 0.0), 2),
            }
            for m in months
        ]
        analysis["totals"] = {
            "cash": round(sum(r["amount"] for r in rows), 2),
            "accrued": round(sum(view_accrual.values()), 2),
            "prepaid": round(sum(view_prepaid.values()), 2),
            "cross_month_packages": len(rows),
            "package_count": len(rows),
        }

    selected_student = db.get(models.Student, student_id_int) if student_id_int else None
    teachers = crud.list_teachers(db)
    return templates.TemplateResponse(
        "finance_payment_detail.html",
        {
            "request": request,
            "start": start_s,
            "end": end_s,
            "coverage_start": coverage_start or "",
            "coverage_end": coverage_end or "",
            "student_id": student_id_int or "",
            "selected_student": selected_student,
            "teacher_id": teacher_id_int or "",
            "teachers": teachers,
            "only_cross_month": (only_cross_month or "").strip() in {"1", "true", "on", "yes"},
            "analysis": analysis,
        },
    )


@app.get("/ui/finance/teacher-pay", response_class=HTMLResponse)
def ui_finance_teacher_pay(
    request: Request,
    start: str | None = None,
    end: str | None = None,
    teacher_id: str | None = None,
    db: Session = Depends(get_db),
):
    require_admin(request)
    start_date, end_date, start_s, end_s = _default_finance_range(start, end)
    teacher_id_int = None
    if teacher_id and str(teacher_id).strip():
        try:
            teacher_id_int = int(str(teacher_id).strip())
        except (ValueError, TypeError):
            teacher_id_int = None
    report = crud.build_teacher_pay_report(
        db,
        start_date=start_date,
        end_date=end_date,
        teacher_id=teacher_id_int,
    )
    teachers = crud.list_teachers(db, active_only=True)
    return templates.TemplateResponse(
        "finance_teacher_pay.html",
        {
            "request": request,
            "start": start_s,
            "end": end_s,
            "teacher_id": teacher_id_int or "",
            "teachers": teachers,
            "report": report,
        },
    )


@app.post("/ui/finance/teacher-pay/rates")
async def ui_finance_teacher_pay_rates(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    form = await request.form()
    updated = 0
    for key, value in form.items():
        if not str(key).startswith("rate_"):
            continue
        try:
            tid = int(str(key).replace("rate_", "", 1))
        except ValueError:
            continue
        raw = str(value).strip().replace(",", ".")
        if raw == "":
            ok = crud.set_teacher_hourly_rate(db, tid, None)
        else:
            try:
                rate = float(raw)
            except ValueError:
                continue
            ok = crud.set_teacher_hourly_rate(db, tid, rate)
        if ok:
            updated += 1
    set_flash_success(request, f"{updated} öğretmen saat ücreti kaydedildi.")
    start = form.get("start") or ""
    end = form.get("end") or ""
    teacher_id = form.get("teacher_id") or ""
    from urllib.parse import urlencode
    qs = urlencode({k: v for k, v in {"start": start, "end": end, "teacher_id": teacher_id}.items() if v})
    return RedirectResponse(
        url="/ui/finance/teacher-pay" + (f"?{qs}" if qs else ""),
        status_code=status.HTTP_303_SEE_OTHER,
    )


@app.get("/ui/finance/expenses", response_class=HTMLResponse)
def ui_finance_expenses(
    request: Request,
    start: str | None = None,
    end: str | None = None,
    category: str | None = None,
    db: Session = Depends(get_db),
):
    require_admin(request)
    start_date, end_date, start_s, end_s = _default_finance_range(start, end)
    cat = (category or "").strip() or None
    items = crud.list_expenses(db, start_date=start_date, end_date=end_date, category=cat)
    total = crud.sum_expenses(db, start_date=start_date, end_date=end_date, category=cat)
    by_category = crud.expense_totals_by_category(db, start_date=start_date, end_date=end_date)
    from datetime import date as date_cls
    return templates.TemplateResponse(
        "finance_expenses.html",
        {
            "request": request,
            "start": start_s,
            "end": end_s,
            "category": cat or "",
            "items": items,
            "total": total,
            "by_category": by_category,
            "categories": crud.EXPENSE_CATEGORIES,
            "today": date_cls.today().isoformat(),
        },
    )


@app.post("/ui/finance/expenses")
def ui_finance_expense_create(
    request: Request,
    title: str = Form(...),
    category: str = Form("Diğer"),
    amount_try: float = Form(...),
    expense_date: str = Form(""),
    method: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    require_admin(request)
    from datetime import date as date_cls
    parsed_date = _parse_optional_date(expense_date) or date_cls.today()
    crud.create_expense(
        db,
        schemas.ExpenseCreate(
            title=title.strip(),
            category=(category or "Diğer").strip() or "Diğer",
            amount_try=amount_try,
            expense_date=parsed_date,
            method=(method or "").strip() or None,
            note=(note or "").strip() or None,
        ),
    )
    set_flash_success(request, "Gider kaydı oluşturuldu.")
    return RedirectResponse(url="/ui/finance/expenses", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/ui/finance/expenses/{expense_id}/delete")
def ui_finance_expense_delete(expense_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    if crud.delete_expense(db, expense_id):
        set_flash_success(request, "Gider kaydı silindi.")
    return RedirectResponse(url="/ui/finance/expenses", status_code=status.HTTP_303_SEE_OTHER)


# UI: Payment Reports
@app.get("/ui/reports/payments", response_class=HTMLResponse)
def payment_reports(request: Request, start: str | None = None, end: str | None = None, course_id: str | None = None, teacher_id: str | None = None, student_id: str | None = None, method: str | None = None, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    if request.session.get("user").get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    from datetime import date
    start_date = None
    end_date = None
    if start:
        try:
            y, m, d = map(int, start.split("-"))
            start_date = date(y, m, d)
        except Exception:
            start_date = None
    if end:
        try:
            y, m, d = map(int, end.split("-"))
            end_date = date(y, m, d)
        except Exception:
            end_date = None
    
    # Query parametrelerini integer'a çevir (boş string'leri None yap)
    course_id_int = None
    teacher_id_int = None
    student_id_int = None
    if course_id and course_id.strip():
        try:
            course_id_int = int(course_id)
        except (ValueError, TypeError):
            course_id_int = None
    if teacher_id and teacher_id.strip():
        try:
            teacher_id_int = int(teacher_id)
        except (ValueError, TypeError):
            teacher_id_int = None
    if student_id and student_id.strip():
        try:
            student_id_int = int(student_id)
        except (ValueError, TypeError):
            student_id_int = None
    
    # Get teacher's students if teacher filter is applied
    teacher_student_ids = None
    if teacher_id_int:
        teacher_students = crud.list_students_by_teacher(db, teacher_id_int)
        if teacher_students:
            teacher_student_ids = [s.id for s in teacher_students]
        else:
            # If teacher has no students, use impossible ID to return no results
            teacher_student_ids = [-1]
    
    # query payments with optional filters and total sum
    q = db.query(models.Payment).join(models.Student)
    # optional joins for filters
    # Filter by course or teacher through enrollments and lessons/payments if needed (basic: by course via enrollments)
    if course_id_int:
        q = q.join(models.Enrollment, models.Enrollment.student_id == models.Payment.student_id).filter(models.Enrollment.course_id == course_id_int)
    if teacher_student_ids is not None:
        # Filter payments by students assigned to the selected teacher
        q = q.filter(models.Payment.student_id.in_(teacher_student_ids))
    if student_id_int:
        # Filter payments by selected student
        q = q.filter(models.Payment.student_id == student_id_int)
    if method and method.strip():
        # Filter payments by payment method
        q = q.filter(models.Payment.method == method.strip())
    if start_date:
        q = q.filter(models.Payment.payment_date >= start_date)
    if end_date:
        q = q.filter(models.Payment.payment_date <= end_date)
    items = q.order_by(models.Payment.payment_date.desc()).all()
    sum_q = db.query(func.coalesce(func.sum(models.Payment.amount_try), 0)).join(models.Student)
    if course_id_int:
        sum_q = sum_q.join(models.Enrollment, models.Enrollment.student_id == models.Payment.student_id).filter(models.Enrollment.course_id == course_id_int)
    if teacher_student_ids is not None:
        # Filter sum by students assigned to the selected teacher
        sum_q = sum_q.filter(models.Payment.student_id.in_(teacher_student_ids))
    if student_id_int:
        # Filter sum by selected student
        sum_q = sum_q.filter(models.Payment.student_id == student_id_int)
    if method and method.strip():
        # Filter sum by payment method
        sum_q = sum_q.filter(models.Payment.method == method.strip())
    if start_date:
        sum_q = sum_q.filter(models.Payment.payment_date >= start_date)
    if end_date:
        sum_q = sum_q.filter(models.Payment.payment_date <= end_date)
    total = float(sum_q.scalar() or 0)
    courses = crud.list_courses(db)
    teachers = crud.list_teachers(db)
    # Get selected student info if student_id is provided
    selected_student = None
    if student_id_int:
        selected_student = db.get(models.Student, student_id_int)
    user = request.session.get("user")
    is_admin = user and user.get("role") == "admin"
    return templates.TemplateResponse("reports_payments.html", {"request": request, "items": items, "total": total, "start": start or "", "end": end or "", "courses": courses, "teachers": teachers, "course_id": course_id or "", "teacher_id": teacher_id or "", "student_id": student_id or "", "method": method or "", "selected_student": selected_student, "is_admin": is_admin})


@app.get("/ui/reports/payments.csv")
def payment_reports_csv(request: Request, start: str | None = None, end: str | None = None, course_id: str | None = None, teacher_id: str | None = None, student_id: str | None = None, method: str | None = None, db: Session = Depends(get_db)):
    if not request.session.get("user"):
        return RedirectResponse(url="/", status_code=302)
    if request.session.get("user").get("role") == "teacher":
        return RedirectResponse(url="/ui/teacher", status_code=302)
    from datetime import date
    import csv
    from io import StringIO
    start_date = None
    end_date = None
    if start:
        try:
            y, m, d = map(int, start.split("-"))
            start_date = date(y, m, d)
        except Exception:
            start_date = None
    if end:
        try:
            y, m, d = map(int, end.split("-"))
            end_date = date(y, m, d)
        except Exception:
            end_date = None
    
    # Query parametrelerini integer'a çevir (boş string'leri None yap)
    course_id_int = None
    teacher_id_int = None
    student_id_int = None
    if course_id and course_id.strip():
        try:
            course_id_int = int(course_id)
        except (ValueError, TypeError):
            course_id_int = None
    if teacher_id and teacher_id.strip():
        try:
            teacher_id_int = int(teacher_id)
        except (ValueError, TypeError):
            teacher_id_int = None
    if student_id and student_id.strip():
        try:
            student_id_int = int(student_id)
        except (ValueError, TypeError):
            student_id_int = None
    
    # Get teacher's students if teacher filter is applied
    teacher_student_ids = None
    if teacher_id_int:
        teacher_students = crud.list_students_by_teacher(db, teacher_id_int)
        if teacher_students:
            teacher_student_ids = [s.id for s in teacher_students]
        else:
            # If teacher has no students, use impossible ID to return no results
            teacher_student_ids = [-1]
    
    q = db.query(models.Payment).join(models.Student)
    if course_id_int:
        q = q.join(models.Enrollment, models.Enrollment.student_id == models.Payment.student_id).filter(models.Enrollment.course_id == course_id_int)
    if teacher_student_ids is not None:
        # Filter payments by students assigned to the selected teacher
        q = q.filter(models.Payment.student_id.in_(teacher_student_ids))
    if student_id_int:
        # Filter payments by selected student
        q = q.filter(models.Payment.student_id == student_id_int)
    if method and method.strip():
        # Filter payments by payment method
        q = q.filter(models.Payment.method == method.strip())
    if start_date:
        q = q.filter(models.Payment.payment_date >= start_date)
    if end_date:
        q = q.filter(models.Payment.payment_date <= end_date)
    items = q.order_by(models.Payment.payment_date.desc()).all()
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Tarih", "Öğrenci", "Tutar", "Yöntem", "Not"])
    for p in items:
        writer.writerow([str(p.payment_date), f"{p.student.first_name} {p.student.last_name}", f"{p.amount_try}", p.method or "", p.note or ""]) 
    return Response(content=buf.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=odeme_raporu.csv"})


# UI: Admin - users
@app.get("/ui/admin/users", response_class=HTMLResponse)
def admin_users(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    users = crud.list_users(db)
    teachers = crud.list_teachers(db)
    return templates.TemplateResponse("admin_users.html", {"request": request, "users": users, "teachers": teachers})


@app.post("/ui/admin/users")
def admin_create_user(request: Request, username: str = Form(...), password: str = Form(...), full_name: str | None = Form(None), role: str | None = Form(None), teacher_id: str | None = Form(None), db: Session = Depends(get_db)):
    require_admin(request)
    try:
        tid = None
        if teacher_id and str(teacher_id).strip():
            try:
                tid = int(str(teacher_id).strip())
            except Exception:
                tid = None
        crud.create_user(db, schemas.UserCreate(username=username, password=password, full_name=full_name, role=role, teacher_id=tid))
    except Exception:
        pass
    return RedirectResponse(url="/ui/admin/users", status_code=302)


@app.post("/ui/admin/users/{user_id}/password")
def admin_change_password(user_id: int, request: Request, password: str = Form(...), db: Session = Depends(get_db)):
    require_admin(request)
    crud.update_user_password(db, user_id, password)
    return RedirectResponse(url="/ui/admin/users", status_code=302)


@app.post("/ui/admin/users/{user_id}/delete")
def admin_delete_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    user = db.get(models.User, user_id)
    if user and user.username != "admin":
        db.delete(user)
        db.commit()
    return RedirectResponse(url="/ui/admin/users", status_code=302)


@app.get("/login/admin", response_class=HTMLResponse)
def login_admin_form(request: Request):
    # Kullanıcı zaten giriş yapmışsa dashboard'a yönlendir
    user = request.session.get("user")
    if user:
        if user.get("role") == "admin":
            return RedirectResponse(url="/dashboard", status_code=302)
        elif user.get("role") == "teacher":
            return RedirectResponse(url="/ui/teacher", status_code=302)
        elif user.get("role") == "staff":
            return RedirectResponse(url="/ui/staff", status_code=302)
    
    # Hata mesajını al
    login_error = request.session.get("login_error", "")
    if login_error:
        request.session.pop("login_error", None)

    return templates.TemplateResponse(
        "login_admin.html",
        {"request": request, "login_error": login_error},
    )


@app.post("/login/admin")
def login_admin(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    try:
        from passlib.hash import pbkdf2_sha256
        user = crud.get_user_by_username(db, username)
        
        # Kullanıcı yoksa
        if not user:
            request.session["login_error"] = "Kullanıcı adı veya şifre hatalı."
            return RedirectResponse(url="/login/admin", status_code=302)
        
        # Şifre kontrolü
        try:
            password_valid = pbkdf2_sha256.verify(password, user.password_hash)
        except Exception as e:
            # Şifre hash hatası
            import logging
            logging.error(f"Şifre doğrulama hatası: {e}")
            request.session["login_error"] = "Giriş hatası. Lütfen tekrar deneyin."
            return RedirectResponse(url="/login/admin", status_code=302)
        
        # Admin kontrolü: role None ise admin kabul et (geriye dönük uyumluluk)
        is_admin = (user.role is None) or (user.role == "admin")
        
        if not password_valid or not is_admin:
            request.session["login_error"] = "Kullanıcı adı veya şifre hatalı, ya da admin yetkisi yok."
            return RedirectResponse(url="/login/admin", status_code=302)
        
        # Session'a kullanıcı bilgilerini kaydet
        request.session["user"] = {
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "role": "admin",
            "teacher_id": getattr(user, 'teacher_id', None),
        }
        # Hata mesajını temizle
        request.session.pop("login_error", None)
        return RedirectResponse(url="/dashboard", status_code=302)
    
    except Exception as e:
        # Genel hata yakalama
        import logging
        import traceback
        logging.error(f"Login hatası: {e}")
        logging.error(traceback.format_exc())
        request.session["login_error"] = f"Sunucu hatası: {str(e)}"
        return RedirectResponse(url="/login/admin", status_code=302)

# Öğretmen için giriş
@app.get("/login/teacher", response_class=HTMLResponse)
def login_teacher_form(request: Request):
    # Kullanıcı zaten giriş yapmışsa ilgili panele yönlendir
    user = request.session.get("user")
    if user:
        if user.get("role") == "teacher":
            return RedirectResponse(url="/ui/teacher", status_code=302)
        elif user.get("role") == "admin":
            return RedirectResponse(url="/dashboard", status_code=302)
        elif user.get("role") == "staff":
            return RedirectResponse(url="/ui/staff", status_code=302)
    
    login_error = request.session.get("login_error", "")
    if login_error:
        request.session.pop("login_error", None)

    return templates.TemplateResponse(
        "login_teacher.html",
        {"request": request, "login_error": login_error},
    )

@app.post("/login/teacher")
def login_teacher(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    try:
        from passlib.hash import pbkdf2_sha256
        user = crud.get_user_by_username(db, username)
        if not user:
            request.session["login_error"] = "Kullanıcı adı veya şifre hatalı."
            return RedirectResponse(url="/login/teacher", status_code=302)
        try:
            password_valid = pbkdf2_sha256.verify(password, user.password_hash)
        except Exception as e:
            import logging
            logging.error(f"Şifre doğrulama hatası: {e}")
            request.session["login_error"] = "Giriş hatası. Lütfen tekrar deneyin."
            return RedirectResponse(url="/login/teacher", status_code=302)
        if not password_valid or user.role != "teacher":
            request.session["login_error"] = "Kullanıcı adı veya şifre hatalı, ya da öğretmen yetkisi yok."
            return RedirectResponse(url="/login/teacher", status_code=302)
        request.session["user"] = {
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "role": "teacher",
            "teacher_id": getattr(user, 'teacher_id', None),
        }
        request.session.pop("login_error", None)
        return RedirectResponse(url="/ui/teacher", status_code=302)
    except Exception as e:
        import logging
        import traceback
        logging.error(f"Login hatası: {e}")
        logging.error(traceback.format_exc())
        request.session["login_error"] = f"Sunucu hatası: {str(e)}"
        return RedirectResponse(url="/login/teacher", status_code=302)

# Personel için giriş (örnek rol adı: staff)
@app.get("/login/staff", response_class=HTMLResponse)
def login_staff_form(request: Request):
    # Kullanıcı zaten giriş yapmışsa ilgili panele yönlendir
    user = request.session.get("user")
    if user:
        if user.get("role") == "staff":
            return RedirectResponse(url="/ui/staff", status_code=302)
        elif user.get("role") == "admin":
            return RedirectResponse(url="/dashboard", status_code=302)
        elif user.get("role") == "teacher":
            return RedirectResponse(url="/ui/teacher", status_code=302)
    
    login_error = request.session.get("login_error", "")
    if login_error:
        request.session.pop("login_error", None)

    return templates.TemplateResponse(
        "login_staff.html",
        {"request": request, "login_error": login_error},
    )

@app.post("/login/staff")
def login_staff(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    try:
        from passlib.hash import pbkdf2_sha256
        user = crud.get_user_by_username(db, username)
        if not user:
            request.session["login_error"] = "Kullanıcı adı veya şifre hatalı."
            return RedirectResponse(url="/login/staff", status_code=302)
        try:
            password_valid = pbkdf2_sha256.verify(password, user.password_hash)
        except Exception as e:
            import logging
            logging.error(f"Şifre doğrulama hatası: {e}")
            request.session["login_error"] = "Giriş hatası. Lütfen tekrar deneyin."
            return RedirectResponse(url="/login/staff", status_code=302)
        if not password_valid or user.role != "staff":
            request.session["login_error"] = "Kullanıcı adı veya şifre hatalı, ya da personel yetkisi yok."
            return RedirectResponse(url="/login/staff", status_code=302)
        request.session["user"] = {
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "role": "staff",
            "teacher_id": getattr(user, 'teacher_id', None),
        }
        request.session.pop("login_error", None)
        return RedirectResponse(url="/ui/staff", status_code=302)
    except Exception as e:
        import logging
        import traceback
        logging.error(f"Login hatası: {e}")
        logging.error(traceback.format_exc())
        request.session["login_error"] = f"Sunucu hatası: {str(e)}"
        return RedirectResponse(url="/login/staff", status_code=302)

@app.get("/ui/staff", response_class=HTMLResponse)
def staff_panel(
	request: Request,
	search: str | None = None,
	student_id: str | None = None,
	teacher_id: str | None = None,
	selected_date: str | None = None,
    attendance_teacher_id: str | None = None,
    attendance_student_id: str | None = None,
    attendance_course_id: str | None = None,
	start_date: str | None = None,
	end_date: str | None = None,
	status: str | None = None,
	order_by: str = "marked_at_desc",
	payment_day_filter: str | None = None,
	payment_status_filter: str | None = None,
	success: str | None = None,
	error: str | None = None,
	db: Session = Depends(get_db),
):
    user = request.session.get("user")
    if not user:
        return RedirectResponse(url="/login/staff", status_code=302)
    # Admin ve staff bu paneli kullanabilir
    if user.get("role") not in ("admin", "staff"):
        if user.get("role") == "teacher":
            return RedirectResponse(url="/ui/teacher", status_code=302)
        return RedirectResponse(url="/login/staff", status_code=302)
    try:
        from sqlalchemy import select
        
        # Query parametrelerini integer'a çevir (boş string'leri None yap)
        student_id_int = None
        teacher_id_int = None
        if student_id and student_id.strip():
            try:
                student_id_int = int(student_id)
            except (ValueError, TypeError):
                student_id_int = None
        if teacher_id and teacher_id.strip():
            try:
                teacher_id_int = int(teacher_id)
            except (ValueError, TypeError):
                teacher_id_int = None
        
        # Tüm öğretmenleri getir
        teachers = crud.list_teachers(db)
        
        # Her öğretmen için haftalık ders programını hazırla
        teachers_schedules = build_teachers_schedules(db, teachers)
        
        # Öğrenci arama ve ders programı
        search_results = []
        selected_student = None
        student_lessons = []
        student_lessons_formatted = []
        selected_student_payments = []
        all_lesson_dates_sorted = []
        student_attendance_summary = {key: 0 for key in crud.ATTENDANCE_STATUS_LABELS}
        attendance_date_entries = []
        
        if search:
            # Öğrenci ara
            search_term = f"%{search.strip()}%"
            students_found = db.query(models.Student).filter(
                (models.Student.first_name.ilike(search_term)) | 
                (models.Student.last_name.ilike(search_term))
            ).limit(20).all()
            # Her öğrenci için ödeme durumunu kontrol et
            search_results = []
            for student in students_found:
                needs_payment = crud.check_student_payment_status(db, student.id)
                search_results.append({
                    "student": student,
                    "needs_payment": needs_payment
                })
        
        if student_id_int:
            # Seçilen öğrencinin bilgilerini ve derslerini getir
            selected_student = crud.get_student(db, student_id_int)
            if selected_student:
                student_lessons = crud.list_lessons_by_student(db, student_id_int)
                # Öğrencinin ödemelerini de getir
                selected_student_payments = crud.list_payments_by_student(db, student_id_int)
                
                # Öğrencinin tüm yoklamalarını tarihe göre sıralı getir (ders tarihleri için)
                student_attendances = db.scalars(
                    select(models.Attendance)
                    .where(models.Attendance.student_id == student_id_int)
                    .order_by(models.Attendance.marked_at.asc())
                ).all()
                student_attendance_summary, attendance_date_entries = crud.summarize_student_attendances(
                    student_attendances
                )
                
                # Dersleri haftalık formata çevir
                from datetime import time as time_type, date as date_type
                # Öğrencinin tüm derslerini birleştir (geçmiş + gelecek)
                # Geçmiş dersler: student_attendances'tan (yoklama alınmış)
                # Gelecek dersler: student_lessons'tan (atanmış)
                
                past_lesson_dates = {entry["date"] for entry in attendance_date_entries}
                
                # Tüm ders tarihlerini birleştir (geçmiş + gelecek)
                # ÖNEMLİ: Toplam ders sayısı için sadece yoklama alınmış dersleri say
                # Gelecek dersler (LessonStudent tablosundan) sadece program gösterimi için kullanılır
                all_lesson_dates = set()
                
                # Geçmiş dersler: Sadece yoklama alınmış dersler
                all_lesson_dates.update(past_lesson_dates)
                
                # Gelecek dersler: LessonStudent tablosundan (program gösterimi için)
                for lesson in student_lessons:
                    all_lesson_dates.add(lesson.lesson_date)
                
                # Tüm tarihleri sırala (program gösterimi için - hem geçmiş hem gelecek)
                all_lesson_dates_sorted = sorted(list(all_lesson_dates))
                
                # Sadece yoklama alınmış derslerin tarihlerini sırala (gösterim için)
                attendance_dates_sorted = [entry["date"] for entry in attendance_date_entries]
                
                # Öğrencinin toplam ders sayısını hesapla (yoklama kayıtları)
                total_lessons_count = sum(student_attendance_summary.values())
                
                # Öğrencinin tüm derslerini tarihe göre sırala (gelecek dersler için)
                all_student_lessons_sorted = sorted(
                    student_lessons,
                    key=lambda x: (x.lesson_date, x.start_time if x.start_time else time_type.min)
                )
                
                for lesson in student_lessons:
                    weekday = WEEKDAY_NAMES[lesson.lesson_date.weekday()] if hasattr(lesson.lesson_date, "weekday") else ""
                    # Dinamik tarih hesapla (bugünden sonraki ilgili gün)
                    current_lesson_date = calculate_next_lesson_date(lesson.lesson_date)
                    
                    # Öğrencinin bu derste toplam dersler içinde kaçıncı ders olduğunu bul
                    # Geçmiş dersler + gelecek dersler birlikte sayılıyor
                    lesson_number = None
                    try:
                        # Bu dersin tarihini bul
                        lesson_date = lesson.lesson_date
                        # Tüm tarihler içinde bu tarihin sırasını bul
                        lesson_index = all_lesson_dates_sorted.index(lesson_date)
                        lesson_number = lesson_index + 1
                    except ValueError:
                        lesson_number = None
                    
                    student_lessons_formatted.append({
                        "weekday": weekday,
                        "lesson": lesson,
                        "current_lesson_date": current_lesson_date,  # Dinamik hesaplanan tarih
                        "lesson_number": lesson_number,
                        "total_same_day": len(all_lesson_dates_sorted)  # Toplam ders sayısı (geçmiş + gelecek)
                    })
            else:
                total_lessons_count = 0
                student_attendances = []
                selected_student_payments = []
        else:
            total_lessons_count = 0
            student_attendances = []
            selected_student_payments = []
        
        # Ödeme durumu tablosu (durum seçilmeden liste oluşturulmaz)
        from datetime import date
        today = date.today()
        payment_status_list = []
        payment_status_filter_value = (payment_status_filter or "").strip().lower()
        if payment_status_filter_value not in crud.VALID_PAYMENT_STATUS_FILTERS:
            payment_status_filter_value = ""
        
        # Geçmişe dönük yoklama için öğretmen ve tarih seçildiğinde öğrencileri getir
        selected_teacher = None
        selected_teacher_lessons = []
        if teacher_id_int and selected_date:
            try:
                import logging
                logging.info(f"🔍 Retrospective attendance: teacher_id={teacher_id_int}, selected_date={selected_date}")
                
                selected_teacher = crud.get_teacher(db, teacher_id_int)
                logging.info(f"✅ Teacher found: {selected_teacher.first_name if selected_teacher else 'None'}")
                
                # Seçilen tarihe ait dersleri getir
                from datetime import datetime
                selected_date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
                selected_weekday = selected_date_obj.weekday()
                logging.info(f"📅 Selected date weekday: {selected_weekday} (0=Mon, 6=Sun)")
                
                # Öğretmene atanmış tüm öğrencileri getir
                teacher_students = db.scalars(
                    select(models.Student)
                    .join(models.TeacherStudent, models.TeacherStudent.student_id == models.Student.id)
                    .where(models.TeacherStudent.teacher_id == teacher_id_int)
                    .order_by(models.Student.first_name.asc(), models.Student.last_name.asc())
                ).all()
                logging.info(f"👥 Total students for teacher: {len(teacher_students)}")
                
                # Öğretmenin o gün hangi dersleri olduğunu bul (haftalık tekrar mantığına göre)
                from sqlalchemy.orm import joinedload
                all_lessons = db.query(models.Lesson).options(
                    joinedload(models.Lesson.course),
                    joinedload(models.Lesson.teacher)
                ).filter(models.Lesson.teacher_id == teacher_id_int).order_by(
                    models.Lesson.lesson_date.asc(),
                    models.Lesson.start_time.asc()
                ).all()
                logging.info(f"📚 Total lessons for teacher: {len(all_lessons)}")
                
                for lesson in all_lessons:
                    lesson_weekday = lesson.lesson_date.weekday()
                    logging.info(f"  - Lesson {lesson.id}: {lesson.course.name}, weekday={lesson_weekday}")
                    
                    # Dersin haftanın hangi günü olduğunu kontrol et
                    if lesson_weekday == selected_weekday:
                        logging.info(f"    ✅ MATCH! Adding lesson {lesson.id} with {len(teacher_students)} students")
                        # Aynı gün içindeki dersler için öğretmene atanmış TÜM öğrencileri ekle
                        selected_teacher_lessons.append({
                            "lesson": lesson,
                            "students": teacher_students  # Öğretmene atanmış tüm öğrenciler
                        })
                    else:
                        logging.info(f"    ❌ NO MATCH: {lesson_weekday} != {selected_weekday}")
                
                logging.info(f"📋 Final selected_teacher_lessons count: {len(selected_teacher_lessons)}")
            except Exception as e:
                import logging
                import traceback
                logging.error(f"❌ Error fetching teacher lessons for date: {e}")
                logging.error(traceback.format_exc())
        
        # Yoklama filtreleme için gerekli verileri hazırla
        students = crud.list_students(db)
        courses = crud.list_courses(db)
        
        # Query parametrelerini integer'a çevir (boş string'leri None yap)
        attendance_teacher_id_int = None
        attendance_student_id_int = None
        attendance_course_id_int = None
        if attendance_teacher_id and attendance_teacher_id.strip():
            try:
                attendance_teacher_id_int = int(attendance_teacher_id)
            except (ValueError, TypeError):
                attendance_teacher_id_int = None
        if attendance_student_id and attendance_student_id.strip():
            try:
                attendance_student_id_int = int(attendance_student_id)
            except (ValueError, TypeError):
                attendance_student_id_int = None
        if attendance_course_id and attendance_course_id.strip():
            try:
                attendance_course_id_int = int(attendance_course_id)
            except (ValueError, TypeError):
                attendance_course_id_int = None
        
        # Tarih filtrelerini parse et
        from datetime import date, datetime
        start_date_obj = None
        end_date_obj = None
        if start_date:
            try:
                y, m, d = map(int, start_date.split("-"))
                start_date_obj = date(y, m, d)
            except Exception:
                pass
        if end_date:
            try:
                y, m, d = map(int, end_date.split("-"))
                end_date_obj = date(y, m, d)
            except Exception:
                pass
        
        # Öğrenci adı filtresi (opsiyonel)
        attendance_student_name = request.query_params.get("attendance_student_name")
        
        # Filtrelerin olup olmadığını kontrol et
        has_filters = any([
            attendance_teacher_id_int is not None,
            attendance_student_id_int is not None,
            attendance_course_id_int is not None,
            status is not None and status.strip(),
            start_date_obj is not None,
            end_date_obj is not None,
            attendance_student_name is not None and attendance_student_name.strip(),
        ])
        
        # Yoklama verilerini filtrele
        attendances = []
        attendances_with_details = []
        if has_filters:
            attendances = crud.list_all_attendances(
                db,
                teacher_id=attendance_teacher_id_int,
                student_id=attendance_student_id_int,
                course_id=attendance_course_id_int,
                status=status,
                start_date=start_date_obj,
                end_date=end_date_obj,
                order_by=order_by,
                limit=200,
            )
            if attendance_student_name and attendance_student_name.strip() and not attendance_student_id_int:
                filtered = []
                for a in attendances:
                    stu = db.get(models.Student, a.student_id)
                    if not stu:
                        continue
                    full_name = f"{stu.first_name} {stu.last_name}"
                    if crud.student_name_matches_prefix(full_name, attendance_student_name):
                        filtered.append(a)
                attendances = filtered

            if attendances:
                lesson_ids = {att.lesson_id for att in attendances}
                student_ids_att = {att.student_id for att in attendances}
                lessons_map = {l.id: l for l in db.scalars(select(models.Lesson).where(models.Lesson.id.in_(lesson_ids))).all()}
                students_map = {s.id: s for s in db.scalars(select(models.Student).where(models.Student.id.in_(student_ids_att))).all()}
                teacher_ids_att = {l.teacher_id for l in lessons_map.values() if l.teacher_id}
                course_ids_att = {l.course_id for l in lessons_map.values() if l.course_id}
                teachers_map = {t.id: t for t in db.scalars(select(models.Teacher).where(models.Teacher.id.in_(teacher_ids_att))).all()} if teacher_ids_att else {}
                courses_map = {c.id: c for c in db.scalars(select(models.Course).where(models.Course.id.in_(course_ids_att))).all()} if course_ids_att else {}
                for att in attendances:
                    lesson = lessons_map.get(att.lesson_id)
                    student = students_map.get(att.student_id)
                    teacher = teachers_map.get(lesson.teacher_id) if lesson and lesson.teacher_id else None
                    course = courses_map.get(lesson.course_id) if lesson and lesson.course_id else None
                    attendances_with_details.append({
                        "attendance": att,
                        "lesson": lesson,
                        "student": student,
                        "teacher": teacher,
                        "course": course,
                    })
        
        # Filtre dict'i oluştur
        filters = {
            "teacher_id": attendance_teacher_id_int,
            "student_id": attendance_student_id_int,
            "course_id": attendance_course_id_int,
            "status": status,
            "start_date": start_date,
            "end_date": end_date,
            "order_by": order_by,
            "student_name": attendance_student_name or "",
        }
        
        return templates.TemplateResponse("staff_panel.html", {
            "request": request,
            "teachers": teachers,
            "teachers_schedules": teachers_schedules,
            "search": search or "",
            "search_results": search_results,
            "selected_student": selected_student,
            "student_lessons": student_lessons_formatted,
            "selected_student_payments": selected_student_payments,
            "total_lessons_count": total_lessons_count if 'total_lessons_count' in locals() else 0,
            "student_attendances": student_attendances if 'student_attendances' in locals() else [],
            "student_attendance_summary": student_attendance_summary if 'student_attendance_summary' in locals() else {},
            "attendance_date_entries": attendance_date_entries if 'attendance_date_entries' in locals() else [],
            "all_lesson_dates_sorted": all_lesson_dates_sorted if 'all_lesson_dates_sorted' in locals() else [],
            "attendance_dates_sorted": attendance_dates_sorted if 'attendance_dates_sorted' in locals() else [],
            "payment_status_list": payment_status_list,
            "payment_day_filter": payment_day_filter or "",
            "payment_status_filter": payment_status_filter_value,
            "payment_day_options": WEEKDAY_NAMES,
            "today": today,
            "selected_teacher": selected_teacher,
            "selected_teacher_id": teacher_id_int,
            "selected_date": selected_date,
            "selected_teacher_lessons": selected_teacher_lessons,
            "success": success,
            "error": error,
            # Yoklama filtreleme için
            "students": students,
            "courses": courses,
            "filters": filters,
            "attendances": attendances_with_details,
        })
    except Exception as e:
        import logging
        logging.error(f"Staff panel template error: {e}")
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html>
        <head><title>Personel Paneli - Piarte</title></head>
        <body>
            <h2>Personel Paneli</h2>
            <p>Hoş geldiniz. Buradan temel işlemleri kolayca erişebilirsiniz:</p>
            <a href="/students/new"><button>Yeni Öğrenci Kaydı</button></a>
            <a href="/lessons/new"><button>Ders Seçimi / Kayıt</button></a>
            <a href="/payments/new"><button>Ödeme Al</button></a>
            <p>Hata: {str(e)}</p>
        </body>
        </html>
        """)

@app.post("/ui/staff/attendance/retrospective")
async def staff_retrospective_attendance(
    request: Request,
    teacher_id: int = Form(...),
    selected_date: str = Form(...),
    db: Session = Depends(get_db)
):
    """Geçmişe dönük yoklama kaydı oluştur"""
    user = request.session.get("user")
    if not user or user.get("role") != "staff":
        return RedirectResponse(url="/login/staff", status_code=302)
    
    try:
        from datetime import datetime
        
        # Form verilerini al
        form_data = await request.form()
        attendance_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        
        # Yoklama kayıtlarını oluştur
        attendance_count = 0
        for key, value in form_data.items():
            if key.startswith("status_"):
                # Format: status_lessonId_studentId
                parts = key.split("_")
                if len(parts) == 3:
                    lesson_id = int(parts[1])
                    student_id = int(parts[2])
                    status_value = value.strip().upper()
                    
                    if status_value:  # Boş değilse
                        # Saat bilgisini al (time_lessonId_studentId formatında)
                        time_key = f"time_{lesson_id}_{student_id}"
                        time_value = form_data.get(time_key, "").strip()
                        
                        # Saat bilgisini parse et
                        marked_at_datetime = None
                        if time_value:
                            try:
                                # time input formatı: "HH:MM"
                                hour, minute = map(int, time_value.split(":"))
                                marked_at_datetime = datetime.combine(attendance_date, datetime.min.time().replace(hour=hour, minute=minute))
                            except (ValueError, AttributeError):
                                # Hata durumunda varsayılan olarak günün başlangıcını kullan
                                marked_at_datetime = datetime.combine(attendance_date, datetime.min.time())
                        else:
                            # Saat girilmemişse günün başlangıcını kullan
                            marked_at_datetime = datetime.combine(attendance_date, datetime.min.time())
                        
                        # Yoklama kaydı oluştur
                        attendance_data = schemas.AttendanceCreate(
                            lesson_id=lesson_id,
                            student_id=student_id,
                            status=status_value,
                            marked_at=marked_at_datetime,
                            note=f"Geçmişe dönük kayıt - {selected_date}"
                        )
                        crud.mark_attendance(db, attendance_data, commit=True)
                        attendance_count += 1
        
        if attendance_count > 0:
            return RedirectResponse(
                url=f"/ui/staff?teacher_id={teacher_id}&selected_date={selected_date}&success={attendance_count} yoklama kaydı başarıyla oluşturuldu",
                status_code=303
            )
        else:
            return RedirectResponse(
                url=f"/ui/staff?teacher_id={teacher_id}&selected_date={selected_date}&error=Hiçbir yoklama durumu seçilmedi",
                status_code=303
            )
    except Exception as e:
        import logging
        logging.error(f"Error creating retrospective attendance: {e}")
        return RedirectResponse(
            url=f"/ui/staff?teacher_id={teacher_id}&selected_date={selected_date}&error=Yoklama kaydı oluşturulurken hata: {str(e)}",
            status_code=303
        )

@app.post("/ui/staff/payment/retrospective")
async def staff_retrospective_payment(
    request: Request,
    student_id: int = Form(...),
    amount: float = Form(...),
    payment_date: str = Form(None),
    note: str = Form(None),
    db: Session = Depends(get_db)
):
    """Staff panelinden ödeme kaydı — tarih her zaman bugün, değiştirilemez."""
    user = request.session.get("user")
    if not user or user.get("role") != "staff":
        return RedirectResponse(url="/login/staff", status_code=302)
    
    try:
        from datetime import date
        
        # Staff panelinde tarih seçilemez; her zaman bugünün tarihi kullanılır
        payment_date_obj = date.today()
        payment_data = schemas.PaymentCreate(
            student_id=student_id,
            amount=amount,
            payment_date=payment_date_obj,
            note=note or f"Staff paneli - {payment_date_obj.isoformat()}"
        )
        crud.create_payment(db, payment_data)
        
        return RedirectResponse(
            url=f"/ui/staff?success=Ödeme kaydı başarıyla oluşturuldu",
            status_code=303
        )
    except Exception as e:
        import logging
        logging.error(f"Error creating retrospective payment: {e}")
        return RedirectResponse(
            url=f"/ui/staff?error=Ödeme kaydı oluşturulurken hata: {str(e)}",
            status_code=303
        )

@app.post("/students/{student_id}/toggle-active")
def toggle_student_active(student_id: int, request: Request, db: Session = Depends(get_db)):
    """Öğrenciyi pasif/aktif yap"""
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "staff"]:
        return RedirectResponse(url="/login/admin", status_code=status.HTTP_303_SEE_OTHER)
    student = db.get(models.Student, student_id)
    if student:
        # Aktif/pasif durumunu tersine çevir
        student.is_active = not student.is_active
        db.commit()
        db.refresh(student)
        status_text = "aktif" if student.is_active else "pasif"
        request.session["student_toggle_success"] = f"Öğrenci {status_text} yapıldı"
    return RedirectResponse(url="/ui/students", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/students/{student_id}/delete")
def delete_student_route(student_id: int, request: Request, db: Session = Depends(get_db)):
    """Öğrenciyi kalıcı olarak siler (admin ve staff)."""
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "staff"]:
        return RedirectResponse(url="/login/admin", status_code=status.HTTP_303_SEE_OTHER)
    if crud.delete_student(db, student_id):
        request.session["student_toggle_success"] = "Öğrenci kalıcı olarak silindi"
    return RedirectResponse(url="/ui/students", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/teachers/{teacher_id}/delete")
def delete_teacher(teacher_id: int, request: Request, db: Session = Depends(get_db)):
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=status.HTTP_303_SEE_OTHER)

    try:
        teacher = crud.delete_teacher(db, teacher_id)
        if teacher:
            name = f"{teacher.first_name} {teacher.last_name}".strip()
            set_flash_success(
                request,
                f"{name} pasif yapıldı. Geçmiş yoklamalar korundu; öğrenciler yeni öğretmene atanabilir.",
            )
        else:
            request.session["flash_error"] = "Öğretmen bulunamadı veya zaten silinmiş."
    except Exception as e:
        db.rollback()
        import logging
        logging.error(f"Öğretmen silme hatası: {e}")
        request.session["flash_error"] = f"Öğretmen silinirken hata oluştu: {str(e)}"

    return RedirectResponse(url="/ui/teachers", status_code=status.HTTP_303_SEE_OTHER)

@app.get("/payments/{payment_id}/edit", response_class=HTMLResponse)
def payment_edit_form(
    payment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    start: str | None = None,
    end: str | None = None,
    course_id: str | None = None,
    teacher_id: str | None = None,
    method: str | None = None,
    return_to: str | None = None,
):
    """Ödeme düzenleme formu (sadece admin için)"""
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=status.HTTP_303_SEE_OTHER)
    
    payment = crud.get_payment(db, payment_id)
    if not payment:
        # Filtre parametrelerini koruyarak geri yönlendir
        params = []
        if start:
            params.append(f"start={start}")
        if end:
            params.append(f"end={end}")
        if course_id:
            params.append(f"course_id={course_id}")
        if teacher_id:
            params.append(f"teacher_id={teacher_id}")
        query_string = "&".join(params)
        redirect_url = f"/ui/reports/payments"
        if query_string:
            redirect_url += "?" + query_string
        request.session["delete_payment_error"] = "Ödeme kaydı bulunamadı."
        return RedirectResponse(url=redirect_url, status_code=status.HTTP_303_SEE_OTHER)
    
    students = crud.list_students(db)
    resolved_return_to = safe_return_url(return_to, "/ui/reports/payments")
    return templates.TemplateResponse("payment_edit.html", {
        "request": request,
        "payment": payment,
        "students": students,
        "start": start or "",
        "end": end or "",
        "course_id": course_id or "",
        "teacher_id": teacher_id or "",
        "method": method or "",
        "return_to": resolved_return_to,
    })


@app.post("/payments/{payment_id}/update")
def update_payment(
    payment_id: int,
    request: Request,
    student_id: int = Form(...),
    amount_try: float = Form(...),
    payment_date: str | None = Form(None),
    method: str | None = Form(None),
    note: str | None = Form(None),
    db: Session = Depends(get_db),
    start: str | None = None,
    end: str | None = None,
    course_id: str | None = None,
    teacher_id: str | None = None,
    method_filter: str | None = None,
    return_to: str | None = Form(None),
):
    """Ödeme kaydını günceller (sadece admin için)"""
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=status.HTTP_303_SEE_OTHER)
    
    from datetime import date
    pd = None
    if payment_date:
        try:
            y, m, d = map(int, payment_date.split("-"))
            pd = date(y, m, d)
        except Exception:
            pd = None
    
    payload = schemas.PaymentUpdate(
        student_id=student_id,
        amount_try=amount_try,
        payment_date=pd,
        method=method,
        note=note,
    )
    
    updated_payment = crud.update_payment(db, payment_id, payload)

    if return_to and str(return_to).strip():
        redirect_url = safe_return_url(return_to, "/ui/reports/payments")
    else:
        params = []
        if start:
            params.append(f"start={start}")
        if end:
            params.append(f"end={end}")
        if course_id:
            params.append(f"course_id={course_id}")
        if teacher_id:
            params.append(f"teacher_id={teacher_id}")
        query_string = "&".join(params)
        redirect_url = "/ui/reports/payments"
        if query_string:
            redirect_url += "?" + query_string

    if updated_payment:
        set_flash_success(request, "Ödeme kaydı başarıyla güncellendi.")
    else:
        request.session["flash_error"] = "Ödeme kaydı güncellenemedi."

    return RedirectResponse(url=redirect_url, status_code=status.HTTP_303_SEE_OTHER)


@app.post("/payments/{payment_id}/delete")
def delete_payment(
    payment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    start: str | None = None,
    end: str | None = None,
    course_id: str | None = None,
    teacher_id: str | None = None,
    method: str | None = None,
    return_to: str | None = None,
):
    """Ödeme kaydını siler (sadece admin için)"""
    user = request.session.get("user")
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/login/admin", status_code=status.HTTP_303_SEE_OTHER)

    success = crud.delete_payment(db, payment_id)

    if return_to and str(return_to).strip():
        redirect_url = safe_return_url(return_to, "/dashboard")
    else:
        params = []
        if start:
            params.append(f"start={start}")
        if end:
            params.append(f"end={end}")
        if course_id:
            params.append(f"course_id={course_id}")
        if teacher_id:
            params.append(f"teacher_id={teacher_id}")
        if method:
            params.append(f"method={method}")
        query_string = "&".join(params)
        redirect_url = "/ui/reports/payments"
        if query_string:
            redirect_url += "?" + query_string

    if success:
        if return_to and str(return_to).strip():
            set_flash_success(request, "Ödeme kaydı başarıyla silindi.")
        else:
            request.session["delete_payment_success"] = "Ödeme kaydı başarıyla silindi."
    else:
        if return_to and str(return_to).strip():
            request.session["flash_error"] = "Ödeme kaydı silinemedi."
        else:
            request.session["delete_payment_error"] = "Ödeme kaydı silinemedi."

    return RedirectResponse(url=redirect_url, status_code=status.HTTP_303_SEE_OTHER)

